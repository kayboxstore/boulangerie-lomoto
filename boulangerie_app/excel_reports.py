from __future__ import annotations

from calendar import monthrange
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from .database import DatabaseHelper
from .report_branding import (
    REPORT_BLUE,
    REPORT_BRAND_NAME_SIZE,
    REPORT_RED,
    REPORT_SUBTITLE_SIZE,
    get_baguette_path,
    get_logo_path,
    get_logo_watermark_path,
)
from .reports import (
    ReportGenerationError,
    get_report_scope_description,
    get_report_scope_label,
    get_report_sections_for_role,
    normalize_role,
    parse_named_amount_lines,
    split_structured_lines,
)
from .status_labels import normalize_status_form_label
from .spreadsheet_security import sanitize_spreadsheet_value
from .version import APP_NAME

TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
THIN_BORDER = Border(
    left=Side(style="thin", color="AEBFD0"),
    right=Side(style="thin", color="AEBFD0"),
    top=Side(style="thin", color="AEBFD0"),
    bottom=Side(style="thin", color="AEBFD0"),
)
TITLE_FONT = Font(name="Poppins", size=16, bold=True, color="FFFFFF")
SECTION_FONT = Font(name="Poppins", size=12, bold=True, color="1F3D5B")
HEADER_FONT = Font(name="Poppins", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Poppins", size=11)
NOTE_FONT = Font(name="Poppins", size=10, italic=True, color="505050")
CASH_BOLD_FONT = Font(name="Poppins", size=11, bold=True)
CASH_GREEN_FONT = Font(name="Poppins", size=11, bold=True, color="1E7D32")
CASH_RED_FONT = Font(name="Poppins", size=11, bold=True, color=REPORT_RED.replace("#", ""))
BRAND_NAME_FONT = Font(
    name="Poppins",
    size=REPORT_BRAND_NAME_SIZE,
    bold=True,
    color=REPORT_RED.replace("#", ""),
)
BRAND_SUBTITLE_FONT = Font(
    name="Poppins",
    size=REPORT_SUBTITLE_SIZE,
    bold=True,
    color=REPORT_BLUE.replace("#", ""),
)
MONEY_FORMAT = '#,##0 "FC"'


def _format_date(target_date: date) -> str:
    return target_date.strftime("%d/%m/%Y")


def _order_accounted_received(row: dict[str, Any]) -> float:
    if "MontantRecuCommande" in row:
        return max(float(row.get("MontantRecuCommande", 0) or 0), 0.0)
    return max(float(row.get("MontantRecu", 0) or 0) - float(row.get("AvanceGeneree", 0) or 0), 0.0)


def _order_gross_received(row: dict[str, Any]) -> float:
    return max(float(row.get("MontantRecu", 0) or 0), 0.0)


def _format_month_label(target_date: date) -> str:
    return target_date.strftime("%m/%Y")


def _month_bounds(target_date: date) -> tuple[date, date]:
    first_day = target_date.replace(day=1)
    last_day = target_date.replace(day=monthrange(target_date.year, target_date.month)[1])
    return first_day, last_day


def _parse_row_date(value: Any) -> date | None:
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return datetime.strptime(text, "%Y-%m-%d").date()
    except ValueError:
        return None


def _filter_rows_for_month(rows: list[dict[str, Any]], date_key: str, reference_date: date) -> list[dict[str, Any]]:
    return [
        row
        for row in rows
        if (row_date := _parse_row_date(row.get(date_key))) is not None
        and row_date.year == reference_date.year
        and row_date.month == reference_date.month
    ]


def _filter_rows_for_period(
    rows: list[dict[str, Any]],
    date_key: str,
    start_date: date,
    end_date: date,
) -> list[dict[str, Any]]:
    return [
        row
        for row in rows
        if (row_date := _parse_row_date(row.get(date_key))) is not None
        and start_date <= row_date <= end_date
    ]


def _normalize_period_bounds(start_date: date, end_date: date) -> tuple[date, date]:
    if end_date < start_date:
        raise ReportGenerationError("La date de fin doit être supérieure ou égale à la date de début.")
    return start_date, end_date


def _format_fc(value: float) -> str:
    return f"{value:,.0f}".replace(",", " ") + " FC"


def _format_number(value: float) -> str:
    if abs(value - int(value)) < 1e-9:
        return str(int(value))
    return f"{value:.2f}".rstrip("0").rstrip(".")


def _safe_text(value: Any) -> str:
    return str(sanitize_spreadsheet_value("" if value is None else str(value)))


def _prevision_quantity(row: dict[str, Any], key: str) -> int:
    return max(int(float(row.get(key, 0) or 0)), 0)


def _prevision_amount(row: dict[str, Any]) -> float:
    return max(float(row.get("MontantPrevu", 0) or 0), 0.0)


def _configure_prevision_printing(sheet: Worksheet, repeat_row: int) -> None:
    sheet.freeze_panes = f"A{repeat_row + 1}"
    sheet.print_title_rows = f"1:{repeat_row}"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.4
    sheet.page_margins.bottom = 0.4


def _write_prevision_order_sheet(
    workbook: Workbook,
    *,
    title: str,
    target_date: date,
    rows: list[dict[str, Any]],
    include_location: bool,
    generated_by: str,
    generated_role: str,
) -> None:
    sheet = workbook.create_sheet(title)
    start_row = _apply_brand_header(
        sheet,
        _format_date(target_date),
        f"Fiche de commande - {title}",
        "Prévisions",
        "Quantités prévues pour préparer la production et les livraisons.",
        generated_by,
        generated_role,
    )
    headers = (["Localisation"] if include_location else []) + [
        "Client",
        "Carré 1.500 FC",
        "Carré 1.000 FC",
        "Baguette 500 FC",
        "Baguette 1.000 FC",
        "Total articles",
        "Montant prévu",
    ]
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(start_row, column_index, header)
        _apply_cell_style(
            cell,
            fill=HEADER_FILL,
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
            font=HEADER_FONT,
        )

    current_row = start_row + 1
    for row in rows:
        values: list[Any] = []
        if include_location:
            values.append(_safe_text(row.get("Localisation", "")))
        values.extend(
            [
                _safe_text(row.get("Client", "")),
                _prevision_quantity(row, "Carre1500"),
                _prevision_quantity(row, "Carre1000"),
                _prevision_quantity(row, "Baguette500"),
                _prevision_quantity(row, "Baguette1000"),
                _prevision_quantity(row, "TotalArticles"),
                _prevision_amount(row),
            ]
        )
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(current_row, column_index, value)
            alignment = Alignment(horizontal="left" if column_index <= (2 if include_location else 1) else "center")
            _apply_cell_style(cell, alignment=alignment)
        sheet.cell(current_row, len(headers)).number_format = MONEY_FORMAT
        current_row += 1

    if rows:
        totals = {
            "Carre1500": sum(_prevision_quantity(row, "Carre1500") for row in rows),
            "Carre1000": sum(_prevision_quantity(row, "Carre1000") for row in rows),
            "Baguette500": sum(_prevision_quantity(row, "Baguette500") for row in rows),
            "Baguette1000": sum(_prevision_quantity(row, "Baguette1000") for row in rows),
            "TotalArticles": sum(_prevision_quantity(row, "TotalArticles") for row in rows),
            "MontantPrevu": sum(_prevision_amount(row) for row in rows),
        }
        label_columns = 2 if include_location else 1
        sheet.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=label_columns,
        )
        sheet.cell(current_row, 1, "TOTAL")
        total_values = [
            totals["Carre1500"],
            totals["Carre1000"],
            totals["Baguette500"],
            totals["Baguette1000"],
            totals["TotalArticles"],
            totals["MontantPrevu"],
        ]
        for column_index in range(1, len(headers) + 1):
            cell = sheet.cell(current_row, column_index)
            if column_index > label_columns:
                cell.value = total_values[column_index - label_columns - 1]
            _apply_cell_style(
                cell,
                fill=SECTION_FILL,
                alignment=Alignment(horizontal="center", vertical="center"),
                font=SECTION_FONT,
            )
        sheet.cell(current_row, len(headers)).number_format = MONEY_FORMAT
    else:
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        sheet.cell(current_row, 1, "Aucune prévision enregistrée pour cette catégorie.")
        _apply_cell_style(
            sheet.cell(current_row, 1),
            alignment=Alignment(horizontal="left"),
            font=NOTE_FONT,
        )

    _configure_prevision_printing(sheet, start_row)
    _autofit_columns(sheet, min_width=13, max_width=28)


def _write_prevision_summary_sheet(
    workbook: Workbook,
    *,
    target_date: date,
    summary: dict[str, Any],
    generated_by: str,
    generated_role: str,
) -> None:
    sheet = workbook.active
    sheet.title = "Résumé"
    start_row = _apply_brand_header(
        sheet,
        _format_date(target_date),
        "Résumé de la prévision de production",
        "Prévisions",
        "Synthèse des commandes prévues et des besoins de production.",
        generated_by,
        generated_role,
    )
    rows = [
        ("Nombre de clients", int(summary.get("NombreClients", 0) or 0), None),
        ("Dépositaires", int(summary.get("NombreDepositaires", 0) or 0), None),
        ("Mamans", int(summary.get("NombreMamans", 0) or 0), None),
        ("Carré 1.500 FC", int(summary.get("TotalCarre1500", 0) or 0), None),
        ("Carré 1.000 FC", int(summary.get("TotalCarre1000", 0) or 0), None),
        ("Baguette 500 FC", int(summary.get("TotalBaguette500", 0) or 0), None),
        ("Baguette 1.000 FC", int(summary.get("TotalBaguette1000", 0) or 0), None),
        ("Total des articles prévus", int(summary.get("TotalArticlesPrevus", 0) or 0), None),
        ("Nombre de sacs à produire", float(summary.get("NombreSacsAProduire", 0) or 0), "0.00"),
        ("Montant prévu", float(summary.get("MontantPrevu", 0) or 0), MONEY_FORMAT),
    ]
    sheet.cell(start_row, 1, "Rubrique")
    sheet.cell(start_row, 2, "Valeur")
    for column_index in (1, 2):
        _apply_cell_style(
            sheet.cell(start_row, column_index),
            fill=HEADER_FILL,
            alignment=Alignment(horizontal="center"),
            font=HEADER_FONT,
        )
    for row_index, (label, value, number_format) in enumerate(rows, start=start_row + 1):
        _apply_cell_style(sheet.cell(row_index, 1, label), alignment=Alignment(horizontal="left"))
        _apply_cell_style(
            sheet.cell(row_index, 2, value),
            alignment=Alignment(horizontal="right"),
            number_format=number_format,
        )
    _configure_prevision_printing(sheet, start_row)
    _autofit_columns(sheet, min_width=18, max_width=38)


def _summarize_production_rows(rows: list[dict[str, Any]]) -> dict[str, float]:
    keys = [
        "NombreBacsCommandes",
        "NombreBacsLivresDepositaires",
        "NombreBacsLivresMamans",
        "NombreBacsDonnes",
        "NombreEchantillons",
        "NombreBacsRestants",
        "NombreBacsFoutus",
        "NombreBacsProduits",
        "NombreSacsUtilises",
    ]
    summary = {key: 0.0 for key in keys}
    for row in rows:
        for key in keys:
            summary[key] += float(row.get(key, 0) or 0)
    ordered = summary["NombreBacsCommandes"]
    produced = summary["NombreBacsProduits"]
    summary["EcartCommandes"] = produced - ordered
    summary["TauxCouverture"] = round((produced * 100.0 / ordered), 2) if ordered > 0 else 0.0
    return summary


def _payroll_total(payrolls: list[dict[str, Any]], key: str) -> float:
    return sum(float(row.get(key, 0) or 0) for row in payrolls)


def _production_field_rows(summary: dict[str, Any]) -> list[tuple[str, Any]]:
    return [
        ("Bacs commandés", float(summary.get("NombreBacsCommandes", 0) or 0)),
        ("Bacs livrés dépositaires", float(summary.get("NombreBacsLivresDepositaires", 0) or 0)),
        ("Bacs livrés mamans", float(summary.get("NombreBacsLivresMamans", 0) or 0)),
        ("Bacs donnés", float(summary.get("NombreBacsDonnes", 0) or 0)),
        ("Échantillons (Agent commercial)", float(summary.get("NombreEchantillons", 0) or 0)),
        ("Bacs restants / disponibles", float(summary.get("NombreBacsRestants", 0) or 0)),
        ("Bacs foutus", float(summary.get("NombreBacsFoutus", 0) or 0)),
        ("Total bacs produits", float(summary.get("NombreBacsProduits", 0) or 0)),
        ("Écart avec commandes", float(summary.get("EcartCommandes", 0) or 0)),
        ("Taux de couverture", f"{_format_number(float(summary.get('TauxCouverture', 0) or 0))} %"),
        ("Nombre de sacs utilisés", float(summary.get("NombreSacsUtilises", 0) or 0)),
        ("Observations", _safe_text(summary.get("Observations")).strip() or "-"),
    ]


def _order_status_summary_rows(rows: list[dict[str, Any]]) -> list[tuple[str, int, int, float, float, float, float, float]]:
    summary: dict[str, dict[str, int | float]] = {}
    for row in rows:
        status = normalize_status_form_label(row.get("Statut")) or "Non précisé"
        bucket = summary.setdefault(
            status,
            {
                "count": 0,
                "trays": 0,
                "expected": 0.0,
                "received": 0.0,
                "advance_used": 0.0,
                "advance_generated": 0.0,
                "debt": 0.0,
            },
        )
        bucket["count"] = int(bucket["count"]) + 1
        bucket["trays"] = int(bucket["trays"]) + int(row.get("NombreBacs", 0) or 0)
        bucket["expected"] = float(bucket["expected"]) + float(row.get("MontantAPercevoir", 0) or 0)
        bucket["received"] = float(bucket["received"]) + _order_accounted_received(row)
        bucket["advance_used"] = float(bucket["advance_used"]) + float(row.get("AvanceUtilisee", 0) or 0)
        bucket["advance_generated"] = float(bucket["advance_generated"]) + float(row.get("AvanceGeneree", 0) or 0)
        bucket["debt"] = float(bucket["debt"]) + float(row.get("Dette", 0) or 0)

    summary_rows: list[tuple[str, int, int, float, float, float, float, float]] = []
    for status in sorted(summary):
        bucket = summary[status]
        summary_rows.append(
            (
                status,
                int(bucket["count"]),
                int(bucket["trays"]),
                float(bucket["expected"]),
                float(bucket["received"]),
                float(bucket["advance_used"]),
                float(bucket["advance_generated"]),
                float(bucket["debt"]),
            )
        )

    if summary_rows:
        summary_rows.append(
            (
                "Total",
                sum(row[1] for row in summary_rows),
                sum(row[2] for row in summary_rows),
                sum(row[3] for row in summary_rows),
                sum(row[4] for row in summary_rows),
                sum(row[5] for row in summary_rows),
                sum(row[6] for row in summary_rows),
                sum(row[7] for row in summary_rows),
            )
        )
    return summary_rows


def _apply_cell_style(
    cell: Any,
    *,
    bold: bool = False,
    fill: PatternFill | None = None,
    alignment: Alignment | None = None,
    number_format: str | None = None,
    font: Font | None = None,
) -> None:
    cell.border = THIN_BORDER
    cell.font = font or Font(name=BODY_FONT.name, size=BODY_FONT.size, bold=bold)
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment
    if number_format is not None:
        cell.number_format = number_format


def _autofit_columns(sheet: Worksheet, min_width: int = 12, max_width: int = 30) -> None:
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, len(value))
        sheet.column_dimensions[letter].width = max(min(length + 2, max_width), min_width)


def _add_table(sheet: Worksheet, start_row: int, end_row: int, end_col: int, name: str) -> None:
    if end_row <= start_row:
        return
    end_col_letter = get_column_letter(end_col)
    table = Table(displayName=name, ref=f"A{start_row}:{end_col_letter}{end_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def _add_brand_image(sheet: Worksheet, image_path: Path, anchor: str, width: int, height: int) -> None:
    if not image_path.exists():
        return
    image = XLImage(str(image_path))
    image.width = width
    image.height = height
    sheet.add_image(image, anchor)


def _add_sheet_watermark(sheet: Worksheet, anchor: str, width: int, height: int) -> None:
    watermark_path = get_logo_watermark_path()
    if not watermark_path.exists():
        watermark_path = get_logo_path()
    if not watermark_path.exists():
        return
    image = XLImage(str(watermark_path))
    image.width = width
    image.height = height
    sheet.add_image(image, anchor)


def _apply_brand_header(
    sheet: Worksheet,
    period_label: str,
    report_title: str,
    scope_label: str,
    scope_description: str,
    generated_by: str,
    generated_role: str,
) -> int:
    sheet.merge_cells("B1:G2")
    title_cell = sheet["B1"]
    title_cell.value = "BOULANGERIE LOMOTO"
    _apply_cell_style(
        title_cell,
        alignment=Alignment(horizontal="center", vertical="center"),
        font=BRAND_NAME_FONT,
    )

    sheet.merge_cells("B3:G4")
    subtitle_cell = sheet["B3"]
    subtitle_cell.value = report_title
    _apply_cell_style(
        subtitle_cell,
        alignment=Alignment(horizontal="center", vertical="center"),
        font=BRAND_SUBTITLE_FONT,
    )

    for row_index, row_height in {
        1: 26,
        2: 26,
        3: 22,
        4: 22,
        5: 16,
        6: 22,
        7: 22,
        8: 22,
        9: 22,
        10: 22,
        11: 22,
    }.items():
        sheet.row_dimensions[row_index].height = row_height

    _add_brand_image(sheet, get_logo_path(), "A1", 56, 56)
    _add_brand_image(sheet, get_baguette_path(), "H1", 84, 32)

    sheet["A6"] = "Période du rapport"
    sheet["B6"] = period_label
    sheet["A7"] = "Profil"
    sheet["B7"] = scope_label
    sheet["A8"] = "Description"
    sheet["B8"] = scope_description
    sheet["A9"] = "Généré le"
    sheet["B9"] = datetime.now().strftime("%d/%m/%Y à %H:%M")
    sheet["A10"] = "Généré par"
    sheet["B10"] = _safe_text(generated_by)
    sheet["A11"] = "Rôle du générateur"
    sheet["B11"] = _safe_text(generated_role)

    for cell_ref in ("A6", "A7", "A8", "A9", "A10", "A11"):
        _apply_cell_style(sheet[cell_ref], bold=True, fill=SECTION_FILL, alignment=Alignment(horizontal="left"))
    for cell_ref in ("B6", "B7", "B8", "B9", "B10", "B11"):
        _apply_cell_style(sheet[cell_ref], alignment=Alignment(horizontal="left", wrap_text=True))

    sheet.column_dimensions["A"].width = 16
    sheet.column_dimensions["B"].width = 24
    for column_letter in ("C", "D", "E", "F", "G"):
        sheet.column_dimensions[column_letter].width = 14
    sheet.column_dimensions["H"].width = 18
    sheet.column_dimensions["I"].width = 10

    return 13


def _apply_cash_emphasis(sheet: Worksheet, row_index: int, label: str, end_column: int = 2) -> None:
    if label in {
        "Reçu commandes",
        "Dettes payées aujourd'hui",
        "Dettes payées du mois",
        "Dettes payées sur la période",
        "Total des entrées",
        "Net à payer des commissions",
    }:
        font = CASH_BOLD_FONT
    elif label in {"Dépenses", "Dépenses du mois", "Dépenses sur la période", "Paies travailleurs", "Total des sorties"}:
        font = CASH_GREEN_FONT
    elif label in {
        "Solde du jour",
        "Solde du mois",
        "Solde sur la période",
        "Solde après paiement des commissions",
        "Solde après paies",
    }:
        font = CASH_RED_FONT
    else:
        return

    for col_index in range(1, end_column + 1):
        sheet.cell(row_index, col_index).font = font


def _build_report_context(
    target_date: date,
    role: str,
    generated_by: str = "",
    generated_role: str | None = None,
) -> dict[str, Any]:
    DatabaseHelper.initialize_database()

    normalized_role = normalize_role(role)
    allowed_sections = get_report_sections_for_role(normalized_role)
    scope_label = get_report_scope_label(normalized_role)
    scope_description = get_report_scope_description(normalized_role)

    stock_journal = DatabaseHelper.get_stock_journal(target_date)
    stock_exits = DatabaseHelper.list_stock_exits_by_date(target_date)
    stock_supplies = DatabaseHelper.list_stock_supplies_by_date(target_date)
    orders = DatabaseHelper.list_orders_by_date(target_date)
    orders_summary = DatabaseHelper.get_orders_summary_for_date(target_date)
    cash = DatabaseHelper.get_cash_for_date(target_date)
    accumulated_debts = DatabaseHelper.get_accumulated_debt_totals_for_date(target_date)
    commissions = DatabaseHelper.list_commissions_by_date(target_date)
    production_summary = DatabaseHelper.get_production_summary_for_date(target_date)
    payrolls = DatabaseHelper.list_payrolls(start_date=target_date, end_date=target_date)
    payroll_summary = DatabaseHelper.get_workers_payroll_summary(start_date=target_date, end_date=target_date)
    expense_details = _safe_text(cash.get("DepensesEffectuees")).strip()
    paid_debts_details = _safe_text(cash.get("DettesPayeesDetails")).strip()
    accumulated_debts_before = float(accumulated_debts.get("DettesAccumuleesAvantPaiement", 0) or 0)
    paid_debts_today = float(cash.get("DettesPayeesAujourdHui", 0) or 0)
    total_payroll_net = _payroll_total(payrolls, "MontantNet")
    total_entries = float(orders_summary.get("MontantRecu", 0) or 0) + paid_debts_today
    balance = total_entries - float(cash.get("MontantTotalDepenses", 0) or 0)
    total_net_commissions = sum(float(row.get("NetAPayer", 0) or 0) for row in commissions)

    return {
        "target_date": target_date,
        "role": normalized_role,
        "allowed_sections": allowed_sections,
        "scope_label": scope_label,
        "scope_description": scope_description,
        "period_label": _format_date(target_date),
        "report_title": f"RAPPORT JOURNALIER - {_format_date(target_date)}",
        "generated_by": generated_by.strip() or "Utilisateur non identifié",
        "generated_role": (generated_role or normalized_role).strip() or normalized_role,
        "stock_journal": stock_journal,
        "stock_exits": stock_exits,
        "stock_supplies": stock_supplies,
        "orders": orders,
        "orders_summary": orders_summary,
        "cash": cash,
        "commissions": commissions,
        "payrolls": payrolls,
        "payroll_summary": payroll_summary,
        "production_summary": production_summary,
        "expense_items": split_structured_lines(expense_details),
        "paid_debts_details": paid_debts_details,
        "paid_debts_items": parse_named_amount_lines(paid_debts_details),
        "total_expected": float(orders_summary.get("MontantAttendu", 0) or 0),
        "total_received": float(orders_summary.get("MontantRecu", 0) or 0),
        "total_received_gross": float(
            orders_summary.get("MontantRecuBrut", orders_summary.get("MontantRecu", 0)) or 0
        ),
        "total_debts": float(orders_summary.get("TotalDettes", 0) or 0),
        "total_trays": int(orders_summary.get("NombreTotalBacs", 0) or 0),
        "total_expenses": float(cash.get("MontantTotalDepenses", 0) or 0),
        "paid_debts_today": paid_debts_today,
        "accumulated_debts_before": accumulated_debts_before,
        "accumulated_debts_remaining": max(accumulated_debts_before - paid_debts_today, 0.0),
        "total_commissions": sum(float(row.get("Commissions", 0) or 0) for row in commissions),
        "total_net_commissions": total_net_commissions,
        "balance_after_commissions": balance - total_net_commissions,
        "total_payroll_net": total_payroll_net,
        "total_entries": total_entries,
        "balance": balance,
        "balance_after_payrolls": balance - total_payroll_net,
    }


def _build_summary_sheet(workbook: Workbook, context: dict[str, Any]) -> None:
    sheet = workbook.active
    sheet.title = "Résumé"
    sheet.freeze_panes = "A13"
    _add_sheet_watermark(sheet, "D7", 260, 260)
    start_row = _apply_brand_header(
        sheet,
        context["period_label"],
        context["report_title"],
        context["scope_label"],
        context["scope_description"],
        context["generated_by"],
        context["generated_role"],
    )
    sheet.cell(start_row, 1, "Indicateur")
    sheet.cell(start_row, 2, "Valeur")
    sheet.cell(start_row, 3, "Type")
    for col in range(1, 4):
        _apply_cell_style(
            sheet.cell(start_row, col),
            fill=HEADER_FILL,
            alignment=Alignment(horizontal="left"),
            font=HEADER_FONT,
        )

    rows: list[tuple[str, int | float, str]] = []
    if "stock" in context["allowed_sections"]:
        rows.extend(
            [
            ]
        )
    if "production" in context["allowed_sections"]:
        production = context["production_summary"]
        rows.extend(
            [
                ("Bacs commandés", float(production.get("NombreBacsCommandes", 0) or 0), "nombre"),
                ("Total bacs produits", float(production.get("NombreBacsProduits", 0) or 0), "nombre"),
                ("Nombre de sacs utilisés", float(production.get("NombreSacsUtilises", 0) or 0), "nombre"),
                ("Écart production", float(production.get("EcartCommandes", 0) or 0), "nombre"),
            ]
        )
    if "orders" in context["allowed_sections"]:
        rows.extend(
            [
                ("Commandes du jour", len(context["orders"]), "nombre"),
                ("Total bacs", context["total_trays"], "nombre"),
                ("Montant attendu", context["total_expected"], "monnaie"),
                ("Payé par clients", context["total_received_gross"], "monnaie"),
                ("Reçu commandes", context["total_received"], "monnaie"),
                ("Dettes", context["total_debts"], "monnaie"),
            ]
        )
    if "cash" in context["allowed_sections"]:
        rows.extend(
            [
                ("Total dettes accumulées", context["accumulated_debts_before"], "monnaie"),
                ("Dettes payées aujourd'hui", context["paid_debts_today"], "monnaie"),
                ("Dettes accumulées restantes", context["accumulated_debts_remaining"], "monnaie"),
                ("Total des entrées", context["total_entries"], "monnaie"),
                ("Dépenses", context["total_expenses"], "monnaie"),
                ("Solde du jour", context["balance"], "monnaie"),
            ]
        )
    if "commissions" in context["allowed_sections"]:
        rows.extend(
            [
                ("Commissions", context["total_commissions"], "monnaie"),
                ("Net à payer des commissions", context["total_net_commissions"], "monnaie"),
                *(
                    [("Solde après paiement des commissions", context["balance_after_commissions"], "monnaie")]
                    if "cash" in context["allowed_sections"]
                    else []
                ),
            ]
        )
    if "workers" in context["allowed_sections"]:
        payroll_summary = context["payroll_summary"]
        rows.extend(
            [
                ("Travailleurs actifs", int(payroll_summary.get("TravailleursActifs", 0) or 0), "nombre"),
                ("Masse salariale mensuelle", float(payroll_summary.get("MasseSalarialeMensuelle", 0) or 0), "monnaie"),
                ("Paies travailleurs", context["total_payroll_net"], "monnaie"),
                ("Solde après paies", context["balance_after_payrolls"], "monnaie"),
            ]
        )

    for row_offset, (label, value, kind) in enumerate(rows, start=1):
        row_index = start_row + row_offset
        sheet.cell(row_index, 1, label)
        value_cell = sheet.cell(row_index, 2, value)
        sheet.cell(row_index, 3, kind)
        _apply_cell_style(sheet.cell(row_index, 1), alignment=Alignment(horizontal="left"))
        _apply_cell_style(sheet.cell(row_index, 2), alignment=Alignment(horizontal="left"))
        _apply_cell_style(sheet.cell(row_index, 3), alignment=Alignment(horizontal="left"))
        if kind == "monnaie":
            value_cell.number_format = MONEY_FORMAT
        _apply_cash_emphasis(sheet, row_index, label, end_column=2)

    end_row = start_row + len(rows)
    _add_table(sheet, start_row, end_row, 3, "ResumeJournalier")
    sheet.column_dimensions["C"].hidden = True

    numeric_rows = [index for index, (_, _, kind) in enumerate(rows, start=1) if kind in {"nombre", "monnaie"}]
    if numeric_rows:
        first_value_row = start_row + numeric_rows[0]
        last_value_row = start_row + numeric_rows[-1]
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "Vue synthétique"
        chart.y_axis.title = "Indicateurs"
        chart.x_axis.title = "Valeurs"
        data = Reference(sheet, min_col=2, min_row=first_value_row, max_row=last_value_row)
        categories = Reference(sheet, min_col=1, min_row=first_value_row, max_row=last_value_row)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(categories)
        chart.height = 7
        chart.width = 12
        sheet.add_chart(chart, "E7")

    _autofit_columns(sheet, min_width=14, max_width=34)


def _build_stock_sheet(workbook: Workbook, context: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("Stock")
    sheet.freeze_panes = "A6"
    _add_sheet_watermark(sheet, "G6", 220, 220)

    sheet["A1"] = "Stock du jour"
    _apply_cell_style(sheet["A1"], fill=TITLE_FILL, alignment=Alignment(horizontal="left"), font=TITLE_FONT)
    sheet["A2"] = f"Date : {_format_date(context['target_date'])}"
    _apply_cell_style(sheet["A2"], alignment=Alignment(horizontal="left"))

    stock_journal = context["stock_journal"]
    row_index = 4
    if stock_journal:
        headers = ["Mouvement", "Farine", "Levure", "Sel", "Huile"]
        values = [
            ["Ouverture", stock_journal.get("FarineOuverture", 0), stock_journal.get("LevureOuverture", 0), stock_journal.get("SelOuverture", 0), stock_journal.get("HuileOuverture", 0)],
            ["Clôture", stock_journal.get("FarineCloture", 0), stock_journal.get("LevureCloture", 0), stock_journal.get("SelCloture", 0), stock_journal.get("HuileCloture", 0)],
        ]
        for col_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row_index, col_index, header)
            _apply_cell_style(cell, fill=HEADER_FILL, alignment=Alignment(horizontal="left"), font=HEADER_FONT)
        for item in values:
            row_index += 1
            for col_index, value in enumerate(item, start=1):
                cell = sheet.cell(row_index, col_index, value)
                _apply_cell_style(cell, alignment=Alignment(horizontal="left"))
        _add_table(sheet, 4, row_index, 5, "StockJournalier")
        for excel_row in range(5, row_index + 1):
            for col in range(2, 6):
                sheet.cell(excel_row, col).number_format = "0.00"
    else:
        sheet["A4"] = "Aucun journal de stock disponible pour cette date."
        _apply_cell_style(sheet["A4"], alignment=Alignment(horizontal="left"), font=NOTE_FONT)

    row_index += 3
    sheet.cell(row_index, 1, "Approvisionnements")
    _apply_cell_style(sheet.cell(row_index, 1), fill=SECTION_FILL, alignment=Alignment(horizontal="left"), font=SECTION_FONT)
    row_index += 1
    supply_headers = ["N°", "Farine", "Levure", "Sel", "Huile"]
    for col_index, header in enumerate(supply_headers, start=1):
        cell = sheet.cell(row_index, col_index, header)
        _apply_cell_style(cell, fill=HEADER_FILL, alignment=Alignment(horizontal="left"), font=HEADER_FONT)
    supplies_start = row_index

    stock_supplies = context["stock_supplies"]
    if stock_supplies:
        for index, item in enumerate(stock_supplies, start=1):
            row_index += 1
            row_values = [
                index,
                item.get("SacsAjoutes", 0),
                item.get("PaquetsAjoutes", 0),
                item.get("KgSelAjoutes", 0),
                item.get("LitresHuileAjoutes", 0),
            ]
            for col_index, value in enumerate(row_values, start=1):
                cell = sheet.cell(row_index, col_index, value)
                _apply_cell_style(cell, alignment=Alignment(horizontal="left"))
            for col in range(2, 6):
                sheet.cell(row_index, col).number_format = "0.00"
        _add_table(sheet, supplies_start, row_index, 5, "ApprovisionnementsStock")
    else:
        row_index += 1
        sheet.cell(row_index, 1, "Aucun approvisionnement enregistré pour cette date.")
        _apply_cell_style(sheet.cell(row_index, 1), alignment=Alignment(horizontal="left"), font=NOTE_FONT)

    row_index += 3
    sheet.cell(row_index, 1, "Sorties de stock")
    _apply_cell_style(sheet.cell(row_index, 1), fill=SECTION_FILL, alignment=Alignment(horizontal="left"), font=SECTION_FONT)
    row_index += 1
    exit_headers = ["N°", "Farine", "Levure", "Sel", "Huile"]
    for col_index, header in enumerate(exit_headers, start=1):
        cell = sheet.cell(row_index, col_index, header)
        _apply_cell_style(cell, fill=HEADER_FILL, alignment=Alignment(horizontal="left"), font=HEADER_FONT)
    exits_start = row_index

    stock_exits = context["stock_exits"]
    if stock_exits:
        for index, item in enumerate(stock_exits, start=1):
            row_index += 1
            row_values = [
                index,
                item.get("SacsUtilises", 0),
                item.get("PaquetsUtilises", 0),
                item.get("KgSelUtilises", 0),
                item.get("LitresHuileUtilises", 0),
            ]
            for col_index, value in enumerate(row_values, start=1):
                cell = sheet.cell(row_index, col_index, value)
                _apply_cell_style(cell, alignment=Alignment(horizontal="left"))
            for col in range(2, 6):
                sheet.cell(row_index, col).number_format = "0.00"
        _add_table(sheet, exits_start, row_index, 5, "SortiesStock")
    else:
        row_index += 1
        sheet.cell(row_index, 1, "Aucune sortie de stock enregistrée pour cette date.")
        _apply_cell_style(sheet.cell(row_index, 1), alignment=Alignment(horizontal="left"), font=NOTE_FONT)

    _autofit_columns(sheet)


def _build_production_sheet(workbook: Workbook, context: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("Production")
    sheet.freeze_panes = "A4"
    _add_sheet_watermark(sheet, "G5", 220, 220)

    sheet["A1"] = "Production du jour"
    _apply_cell_style(sheet["A1"], fill=TITLE_FILL, alignment=Alignment(horizontal="left"), font=TITLE_FONT)
    sheet["A2"] = f"Date : {_format_date(context['target_date'])}"
    _apply_cell_style(sheet["A2"], alignment=Alignment(horizontal="left"))

    headers = ["Champ", "Valeur"]
    for col_index, header in enumerate(headers, start=1):
        _apply_cell_style(sheet.cell(4, col_index, header), fill=HEADER_FILL, alignment=Alignment(horizontal="left"), font=HEADER_FONT)

    current_row = 4
    for label, value in _production_field_rows(context["production_summary"]):
        current_row += 1
        _apply_cell_style(sheet.cell(current_row, 1, label), alignment=Alignment(horizontal="left"))
        _apply_cell_style(sheet.cell(current_row, 2, value), alignment=Alignment(horizontal="left", wrap_text=True))
    _add_table(sheet, 4, current_row, 2, "ProductionJournaliere")
    _autofit_columns(sheet)


def _build_orders_sheet(workbook: Workbook, context: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("Commandes")
    sheet.freeze_panes = "A4"
    _add_sheet_watermark(sheet, "G5", 220, 220)

    sheet["A1"] = "Commandes du jour"
    _apply_cell_style(sheet["A1"], fill=TITLE_FILL, alignment=Alignment(horizontal="left"), font=TITLE_FONT)
    sheet["A2"] = f"Date : {_format_date(context['target_date'])}"
    _apply_cell_style(sheet["A2"], alignment=Alignment(horizontal="left"))

    headers = ["Client", "Statut", "Bacs", "À percevoir", "Payé client", "Reçu commande", "Avance utilisée", "Nouvelle avance", "Solde avance", "Dette"]
    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(3, col_index, header)
        _apply_cell_style(cell, fill=HEADER_FILL, alignment=Alignment(horizontal="left"), font=HEADER_FONT)

    orders = context["orders"]
    current_row = 3
    if orders:
        for item in orders:
            current_row += 1
            values = [
                _safe_text(item.get("Client")),
                normalize_status_form_label(item.get("Statut")),
                int(item.get("NombreBacs", 0) or 0),
                float(item.get("MontantAPercevoir", 0) or 0),
                _order_gross_received(item),
                _order_accounted_received(item),
                float(item.get("AvanceUtilisee", 0) or 0),
                float(item.get("AvanceGeneree", 0) or 0),
                float(item.get("SoldeAvance", 0) or 0),
                float(item.get("Dette", 0) or 0),
            ]
            for col_index, value in enumerate(values, start=1):
                cell = sheet.cell(current_row, col_index, value)
                _apply_cell_style(cell, alignment=Alignment(horizontal="left"))
        for col in range(4, 11):
            sheet.cell(current_row, col).number_format = MONEY_FORMAT

        totals_row = current_row + 1
        sheet.cell(totals_row, 1, "Totaux")
        sheet.cell(totals_row, 3, f"=SUM(C4:C{current_row})")
        sheet.cell(totals_row, 4, f"=SUM(D4:D{current_row})")
        sheet.cell(totals_row, 5, f"=SUM(E4:E{current_row})")
        sheet.cell(totals_row, 6, f"=SUM(F4:F{current_row})")
        sheet.cell(totals_row, 7, f"=SUM(G4:G{current_row})")
        sheet.cell(totals_row, 8, f"=SUM(H4:H{current_row})")
        sheet.cell(totals_row, 10, f"=SUM(J4:J{current_row})")
        for col_index in range(1, 11):
            _apply_cell_style(
                sheet.cell(totals_row, col_index),
                fill=SECTION_FILL,
                alignment=Alignment(horizontal="left"),
                font=SECTION_FONT if col_index == 1 else BODY_FONT,
            )
        for col in range(4, 11):
            sheet.cell(totals_row, col).number_format = MONEY_FORMAT
        _add_table(sheet, 3, current_row, 10, "CommandesJour")
    else:
        sheet["A4"] = "Aucune commande enregistrée pour cette date."
        _apply_cell_style(sheet["A4"], alignment=Alignment(horizontal="left"), font=NOTE_FONT)

    _autofit_columns(sheet)


def _build_cash_sheet(workbook: Workbook, context: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("Caisse")
    sheet.freeze_panes = "A4"
    _add_sheet_watermark(sheet, "G5", 220, 220)

    sheet["A1"] = "Caisse du jour"
    _apply_cell_style(sheet["A1"], fill=TITLE_FILL, alignment=Alignment(horizontal="left"), font=TITLE_FONT)
    sheet["A2"] = f"Date : {_format_date(context['target_date'])}"
    _apply_cell_style(sheet["A2"], alignment=Alignment(horizontal="left"))

    headers = ["Champ", "Valeur"]
    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(4, col_index, header)
        _apply_cell_style(cell, fill=HEADER_FILL, alignment=Alignment(horizontal="left"), font=HEADER_FONT)

    rows = [
        ("Montant attendu", context["total_expected"]),
        ("Payé par clients", context["total_received_gross"]),
        ("Reçu commandes", context["total_received"]),
        ("Dettes", context["total_debts"]),
        ("Total dettes accumulées", context["accumulated_debts_before"]),
        ("Dettes payées aujourd'hui", context["paid_debts_today"]),
        ("Dettes accumulées restantes", context["accumulated_debts_remaining"]),
        ("Total des entrées", context["total_entries"]),
        ("Dépenses", context["total_expenses"]),
        *(
            [
                ("Paies travailleurs", context["total_payroll_net"]),
                ("Total des sorties", context["total_expenses"] + context["total_payroll_net"]),
                ("Solde après paies", context["balance_after_payrolls"]),
            ]
            if "workers" in context["allowed_sections"]
            else []
        ),
        ("Solde du jour", context["balance"]),
    ]
    current_row = 4
    for label, value in rows:
        current_row += 1
        sheet.cell(current_row, 1, label)
        sheet.cell(current_row, 2, value)
        _apply_cell_style(sheet.cell(current_row, 1), alignment=Alignment(horizontal="left"))
        _apply_cell_style(sheet.cell(current_row, 2), alignment=Alignment(horizontal="left"), number_format=MONEY_FORMAT)
        _apply_cash_emphasis(sheet, current_row, label, end_column=2)
    _add_table(sheet, 4, current_row, 2, "CaisseJour")

    current_row += 3
    sheet.cell(current_row, 1, "Liste des dépenses")
    _apply_cell_style(sheet.cell(current_row, 1), fill=SECTION_FILL, alignment=Alignment(horizontal="left"), font=SECTION_FONT)
    current_row += 1
    expense_items = context["expense_items"]
    if expense_items:
        expense_header_row = current_row
        for col_index, header in enumerate(("N°", "Détail"), start=1):
            cell = sheet.cell(expense_header_row, col_index, header)
            _apply_cell_style(cell, fill=HEADER_FILL, alignment=Alignment(horizontal="left"), font=HEADER_FONT)
        for index, item in enumerate(expense_items, start=1):
            current_row += 1
            sheet.cell(current_row, 1, index)
            sheet.cell(current_row, 2, item)
            _apply_cell_style(sheet.cell(current_row, 1), alignment=Alignment(horizontal="left"))
            _apply_cell_style(sheet.cell(current_row, 2), alignment=Alignment(horizontal="left", wrap_text=True))
        _add_table(sheet, expense_header_row, current_row, 2, "ListeDepensesJour")
    else:
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 1, end_column=4)
        details_cell = sheet.cell(current_row, 1, "Aucune dépense détaillée n'a été enregistrée pour cette date.")
        _apply_cell_style(details_cell, alignment=Alignment(horizontal="left", vertical="top", wrap_text=True), font=NOTE_FONT)
        current_row += 1

    current_row += 3
    sheet.cell(current_row, 1, "Ceux qui ont payé leurs dettes")
    _apply_cell_style(sheet.cell(current_row, 1), fill=SECTION_FILL, alignment=Alignment(horizontal="left"), font=SECTION_FONT)
    current_row += 1
    paid_debts_items = context["paid_debts_items"]
    if paid_debts_items:
        has_amounts = any(amount for _, amount in paid_debts_items)
        headers = ("Nom", "Montant payé") if has_amounts else ("N°", "Personne")
        paid_header_row = current_row
        for col_index, header in enumerate(headers, start=1):
            cell = sheet.cell(paid_header_row, col_index, header)
            _apply_cell_style(cell, fill=HEADER_FILL, alignment=Alignment(horizontal="left"), font=HEADER_FONT)
        for index, (name, amount) in enumerate(paid_debts_items, start=1):
            current_row += 1
            if has_amounts:
                values = (name, amount or "-")
            else:
                values = (index, name)
            for col_index, value in enumerate(values, start=1):
                cell = sheet.cell(current_row, col_index, value)
                _apply_cell_style(cell, alignment=Alignment(horizontal="left", wrap_text=True))
        _add_table(sheet, paid_header_row, current_row, 2, "DettesPayeesJour")
    else:
        note = (
            "Aucune personne n'a été renseignée dans la liste des dettes payées pour cette date."
            if float(context["paid_debts_today"] or 0) > 0
            else "Aucun paiement de dette n'a été enregistré pour cette date."
        )
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 1, end_column=4)
        paid_note_cell = sheet.cell(current_row, 1, note)
        _apply_cell_style(paid_note_cell, alignment=Alignment(horizontal="left", vertical="top", wrap_text=True), font=NOTE_FONT)
        current_row += 1

    current_row += 3
    sheet.cell(current_row, 1, "NB")
    _apply_cell_style(sheet.cell(current_row, 1), fill=SECTION_FILL, alignment=Alignment(horizontal="left"), font=SECTION_FONT)
    current_row += 1
    payroll_sentence = ""
    if "workers" in context["allowed_sections"]:
        payroll_sentence = (
            f" Les paies des travailleurs représentent {_format_fc(context['total_payroll_net'])}; "
            f"le solde réel après paies est de {_format_fc(context['balance_after_payrolls'])}."
        )
    recap_text = (
        f"Les entrées du jour correspondent au reçu commandes ({context['total_received']:,.0f} FC) "
        f"additionné aux dettes payées aujourd'hui ({context['paid_debts_today']:,.0f} FC), "
        f"soit un total des entrées de {context['total_entries']:,.0f} FC. "
        f"Les sorties du jour correspondent aux dépenses enregistrées, soit {context['total_expenses']:,.0f} FC. "
        f"Le solde du jour ressort donc à {context['balance']:,.0f} FC. "
        f"Les dettes accumulées restantes sont de {context['accumulated_debts_remaining']:,.0f} FC."
        f"{payroll_sentence}"
    ).replace(",", " ")
    sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 2, end_column=5)
    recap_cell = sheet.cell(current_row, 1, recap_text)
    _apply_cell_style(recap_cell, alignment=Alignment(horizontal="left", vertical="top", wrap_text=True))

    _autofit_columns(sheet)
    sheet.column_dimensions["D"].width = 18


def _build_commissions_sheet(workbook: Workbook, context: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("Commissions")
    sheet.freeze_panes = "A4"
    _add_sheet_watermark(sheet, "G5", 220, 220)

    sheet["A1"] = "Commissions du jour"
    _apply_cell_style(sheet["A1"], fill=TITLE_FILL, alignment=Alignment(horizontal="left"), font=TITLE_FONT)
    sheet["A2"] = f"Date : {_format_date(context['target_date'])}"
    _apply_cell_style(sheet["A2"], alignment=Alignment(horizontal="left"))

    headers = ["Nom", "Statut", "Bacs", "Payé", "Commission", "Dette", "Net"]
    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(3, col_index, header)
        _apply_cell_style(cell, fill=HEADER_FILL, alignment=Alignment(horizontal="left"), font=HEADER_FONT)

    commissions = context["commissions"]
    current_row = 3
    if commissions:
        for item in commissions:
            current_row += 1
            values = [
                _safe_text(item.get("Nom")),
                normalize_status_form_label(item.get("Statut")),
                int(item.get("NombreBacs", 0) or 0),
                float(item.get("MontantPaye", 0) or 0),
                float(item.get("Commissions", 0) or 0),
                float(item.get("Dettes", 0) or 0),
                float(item.get("NetAPayer", 0) or 0),
            ]
            for col_index, value in enumerate(values, start=1):
                cell = sheet.cell(current_row, col_index, value)
                _apply_cell_style(cell, alignment=Alignment(horizontal="left"))
            for col in range(4, 8):
                sheet.cell(current_row, col).number_format = MONEY_FORMAT

        totals_row = current_row + 1
        sheet.cell(totals_row, 1, "Totaux")
        sheet.cell(totals_row, 3, f"=SUM(C4:C{current_row})")
        sheet.cell(totals_row, 4, f"=SUM(D4:D{current_row})")
        sheet.cell(totals_row, 5, f"=SUM(E4:E{current_row})")
        sheet.cell(totals_row, 6, f"=SUM(F4:F{current_row})")
        sheet.cell(totals_row, 7, f"=SUM(G4:G{current_row})")
        for col_index in range(1, 8):
            _apply_cell_style(
                sheet.cell(totals_row, col_index),
                fill=SECTION_FILL,
                alignment=Alignment(horizontal="left"),
                font=SECTION_FONT if col_index == 1 else BODY_FONT,
            )
        for col in range(4, 8):
            sheet.cell(totals_row, col).number_format = MONEY_FORMAT
        _add_table(sheet, 3, current_row, 7, "CommissionsJour")
    else:
        sheet["A4"] = "Aucune commission enregistrée pour cette date."
        _apply_cell_style(sheet["A4"], alignment=Alignment(horizontal="left"), font=NOTE_FONT)

    _autofit_columns(sheet)


def _build_payroll_sheet(
    workbook: Workbook,
    context: dict[str, Any],
    title: str,
    table_name: str,
    *,
    include_date: bool = True,
) -> None:
    sheet = workbook.create_sheet("Travailleurs et paies")
    sheet.freeze_panes = "A4"
    _add_sheet_watermark(sheet, "G5", 220, 220)

    sheet["A1"] = title
    _apply_cell_style(sheet["A1"], fill=TITLE_FILL, alignment=Alignment(horizontal="left"), font=TITLE_FONT)
    sheet["A2"] = f"Période : {context['period_label']}"
    _apply_cell_style(sheet["A2"], alignment=Alignment(horizontal="left"))

    payroll_summary = context["payroll_summary"]
    summary_rows = [
        ("Travailleurs enregistrés", int(payroll_summary.get("NombreTravailleurs", 0) or 0), "nombre"),
        ("Travailleurs actifs", int(payroll_summary.get("TravailleursActifs", 0) or 0), "nombre"),
        ("Masse salariale mensuelle", float(payroll_summary.get("MasseSalarialeMensuelle", 0) or 0), "monnaie"),
        ("Paies enregistrées", len(context["payrolls"]), "nombre"),
        ("Montant brut", _payroll_total(context["payrolls"], "MontantBrut"), "monnaie"),
        ("Primes", _payroll_total(context["payrolls"], "Prime"), "monnaie"),
        ("Avances", _payroll_total(context["payrolls"], "Avance"), "monnaie"),
        ("Retenues", _payroll_total(context["payrolls"], "Retenue"), "monnaie"),
        ("Net payé", context["total_payroll_net"], "monnaie"),
    ]
    sheet.cell(4, 1, "Indicateur")
    sheet.cell(4, 2, "Valeur")
    for col in range(1, 3):
        _apply_cell_style(sheet.cell(4, col), fill=HEADER_FILL, alignment=Alignment(horizontal="left"), font=HEADER_FONT)
    for row_index, (label, value, kind) in enumerate(summary_rows, start=5):
        _apply_cell_style(sheet.cell(row_index, 1, label), alignment=Alignment(horizontal="left"))
        value_cell = sheet.cell(row_index, 2, value)
        _apply_cell_style(value_cell, alignment=Alignment(horizontal="left"))
        if kind == "monnaie":
            value_cell.number_format = MONEY_FORMAT
        _apply_cash_emphasis(sheet, row_index, label, end_column=2)
    _add_table(sheet, 4, 4 + len(summary_rows), 2, f"{table_name}Resume")

    current_row = 4 + len(summary_rows) + 3
    sheet.cell(current_row, 1, "Détail des paies")
    _apply_cell_style(sheet.cell(current_row, 1), fill=SECTION_FILL, alignment=Alignment(horizontal="left"), font=SECTION_FONT)
    current_row += 1
    headers = ["Travailleur", "Fonction", "Période", "Brut", "Prime", "Avance", "Retenue", "Net", "Mode", "Statut"]
    if include_date:
        headers.insert(0, "Date")
    for col_index, header in enumerate(headers, start=1):
        _apply_cell_style(sheet.cell(current_row, col_index, header), fill=HEADER_FILL, alignment=Alignment(horizontal="left"), font=HEADER_FONT)
    detail_start = current_row

    if context["payrolls"]:
        for item in context["payrolls"]:
            current_row += 1
            values: list[Any] = []
            if include_date:
                row_date = _parse_row_date(item.get("DatePaie"))
                values.append(_format_date(row_date) if row_date is not None else _safe_text(item.get("DatePaie")))
            values.extend(
                [
                    _safe_text(item.get("NomComplet")),
                    _safe_text(item.get("Fonction")),
                    _safe_text(item.get("Periode")),
                    float(item.get("MontantBrut", 0) or 0),
                    float(item.get("Prime", 0) or 0),
                    float(item.get("Avance", 0) or 0),
                    float(item.get("Retenue", 0) or 0),
                    float(item.get("MontantNet", 0) or 0),
                    _safe_text(item.get("ModePaiement")),
                    _safe_text(item.get("Statut")),
                ]
            )
            for col_index, value in enumerate(values, start=1):
                _apply_cell_style(sheet.cell(current_row, col_index, value), alignment=Alignment(horizontal="left"))
            first_money_col = 5 if include_date else 4
            for col in range(first_money_col, first_money_col + 5):
                sheet.cell(current_row, col).number_format = MONEY_FORMAT

        totals_row = current_row + 1
        sheet.cell(totals_row, 1, "Totaux")
        first_money_col = 5 if include_date else 4
        for offset in range(5):
            col_index = first_money_col + offset
            col_letter = get_column_letter(col_index)
            sheet.cell(totals_row, col_index, f"=SUM({col_letter}{detail_start + 1}:{col_letter}{current_row})")
            sheet.cell(totals_row, col_index).number_format = MONEY_FORMAT
        for col_index in range(1, len(headers) + 1):
            _apply_cell_style(
                sheet.cell(totals_row, col_index),
                fill=SECTION_FILL,
                alignment=Alignment(horizontal="left"),
                font=SECTION_FONT if col_index == 1 else BODY_FONT,
            )
        _add_table(sheet, detail_start, current_row, len(headers), table_name)
    else:
        current_row += 1
        _apply_cell_style(
            sheet.cell(current_row, 1, "Aucune paie de travailleur enregistrée pour cette période."),
            alignment=Alignment(horizontal="left"),
            font=NOTE_FONT,
        )

    _autofit_columns(sheet, min_width=13, max_width=32)


def _build_monthly_report_context(
    target_date: date,
    role: str,
    generated_by: str = "",
    generated_role: str | None = None,
) -> dict[str, Any]:
    DatabaseHelper.initialize_database()

    normalized_role = normalize_role(role)
    allowed_sections = get_report_sections_for_role(normalized_role)
    scope_label = get_report_scope_label(normalized_role)
    scope_description = get_report_scope_description(normalized_role)
    first_day, last_day = _month_bounds(target_date)
    month_label = _format_month_label(target_date)

    stock_exits = _filter_rows_for_month(DatabaseHelper.list_stock_exits(), "DateSortie", target_date)
    stock_supplies = _filter_rows_for_month(DatabaseHelper.list_stock_supplies(), "DateApprovisionnement", target_date)
    stock_journals: list[dict[str, Any]] = []
    for day_number in range(1, last_day.day + 1):
        current_day = target_date.replace(day=day_number)
        journal = DatabaseHelper.get_stock_journal(current_day)
        if journal:
            stock_journals.append(journal)

    orders = _filter_rows_for_month(DatabaseHelper.list_orders(), "DateCommande", target_date)
    cash_days = DatabaseHelper.list_cash_balance_by_period(first_day, last_day)
    commissions = _filter_rows_for_month(DatabaseHelper.list_commissions(), "DateCommission", target_date)
    productions = _filter_rows_for_month(DatabaseHelper.list_productions(), "DateProduction", target_date)
    payrolls = DatabaseHelper.list_payrolls(start_date=first_day, end_date=last_day)
    payroll_summary = DatabaseHelper.get_workers_payroll_summary(start_date=first_day, end_date=last_day)
    production_summary = _summarize_production_rows(productions)

    expense_items_by_day: list[tuple[str, str]] = []
    paid_debts_items_by_day: list[tuple[str, str, str]] = []
    for row in cash_days:
        row_date = _parse_row_date(row.get("DateCaisse"))
        row_date_label = _format_date(row_date) if row_date is not None else _safe_text(row.get("DateCaisse"))
        for item in split_structured_lines(_safe_text(row.get("DepensesEffectuees")).strip()):
            expense_items_by_day.append((row_date_label, item))
        for name, amount in parse_named_amount_lines(_safe_text(row.get("DettesPayeesDetails")).strip()):
            paid_debts_items_by_day.append((row_date_label, name, amount or "-"))

    total_trays = sum(int(row.get("NombreBacs", 0) or 0) for row in orders)
    total_expected = sum(float(row.get("MontantAPercevoir", 0) or 0) for row in orders)
    total_received = sum(_order_accounted_received(row) for row in orders)
    total_received_gross = sum(_order_gross_received(row) for row in orders)
    total_debts = sum(float(row.get("Dette", 0) or 0) for row in orders)
    total_expenses = sum(float(row.get("MontantTotalDepenses", 0) or 0) for row in cash_days)
    paid_debts_month = sum(float(row.get("DettesPayeesAujourdHui", 0) or 0) for row in cash_days)
    total_entries = total_received + paid_debts_month
    balance = total_entries - total_expenses
    total_commissions = sum(float(row.get("Commissions", 0) or 0) for row in commissions)
    total_net_commissions = sum(float(row.get("NetAPayer", 0) or 0) for row in commissions)
    total_payroll_net = _payroll_total(payrolls, "MontantNet")
    balance_after_commissions = balance - total_net_commissions
    remaining_accumulated_debts = float(cash_days[-1].get("DettesAccumuleesRestantes", 0) or 0) if cash_days else 0.0

    return {
        "target_date": target_date,
        "month_label": month_label,
        "first_day": first_day,
        "last_day": last_day,
        "period_label": f"Du {_format_date(first_day)} au {_format_date(last_day)}",
        "report_title": f"RAPPORT MENSUEL - {month_label}",
        "role": normalized_role,
        "allowed_sections": allowed_sections,
        "scope_label": scope_label,
        "scope_description": scope_description,
        "generated_by": generated_by.strip() or "Utilisateur non identifié",
        "generated_role": (generated_role or normalized_role).strip() or normalized_role,
        "stock_journals": stock_journals,
        "stock_exits": stock_exits,
        "stock_supplies": stock_supplies,
        "orders": orders,
        "cash_days": cash_days,
        "commissions": commissions,
        "payrolls": payrolls,
        "payroll_summary": payroll_summary,
        "productions": productions,
        "production_summary": production_summary,
        "expense_items_by_day": expense_items_by_day,
        "paid_debts_items_by_day": paid_debts_items_by_day,
        "total_expected": total_expected,
        "total_received": total_received,
        "total_received_gross": total_received_gross,
        "total_debts": total_debts,
        "total_trays": total_trays,
        "total_expenses": total_expenses,
        "paid_debts_month": paid_debts_month,
        "total_entries": total_entries,
        "balance": balance,
        "remaining_accumulated_debts": remaining_accumulated_debts,
        "total_commissions": total_commissions,
        "total_net_commissions": total_net_commissions,
        "balance_after_commissions": balance_after_commissions,
        "total_payroll_net": total_payroll_net,
        "balance_after_payrolls": balance - total_payroll_net,
        "total_farine": sum(float(row.get("SacsUtilises", 0) or 0) for row in stock_exits),
        "total_levure": sum(float(row.get("PaquetsUtilises", 0) or 0) for row in stock_exits),
        "total_sel": sum(float(row.get("KgSelUtilises", 0) or 0) for row in stock_exits),
        "total_huile": sum(float(row.get("LitresHuileUtilises", 0) or 0) for row in stock_exits),
        "total_farine_added": sum(float(row.get("SacsAjoutes", 0) or 0) for row in stock_supplies),
        "total_levure_added": sum(float(row.get("PaquetsAjoutes", 0) or 0) for row in stock_supplies),
        "total_sel_added": sum(float(row.get("KgSelAjoutes", 0) or 0) for row in stock_supplies),
        "total_huile_added": sum(float(row.get("LitresHuileAjoutes", 0) or 0) for row in stock_supplies),
    }


def _build_period_report_context(
    start_date: date,
    end_date: date,
    role: str,
    generated_by: str = "",
    generated_role: str | None = None,
) -> dict[str, Any]:
    DatabaseHelper.initialize_database()

    start_date, end_date = _normalize_period_bounds(start_date, end_date)
    normalized_role = normalize_role(role)
    allowed_sections = get_report_sections_for_role(normalized_role)
    scope_label = get_report_scope_label(normalized_role)
    scope_description = get_report_scope_description(normalized_role)

    stock_exits = _filter_rows_for_period(DatabaseHelper.list_stock_exits(), "DateSortie", start_date, end_date)
    stock_supplies = _filter_rows_for_period(
        DatabaseHelper.list_stock_supplies(),
        "DateApprovisionnement",
        start_date,
        end_date,
    )
    stock_journals: list[dict[str, Any]] = []
    for day_offset in range((end_date - start_date).days + 1):
        current_day = date.fromordinal(start_date.toordinal() + day_offset)
        journal = DatabaseHelper.get_stock_journal(current_day)
        if journal:
            stock_journals.append(journal)

    orders = _filter_rows_for_period(DatabaseHelper.list_orders(), "DateCommande", start_date, end_date)
    cash_days = DatabaseHelper.list_cash_balance_by_period(start_date, end_date)
    commissions = _filter_rows_for_period(DatabaseHelper.list_commissions(), "DateCommission", start_date, end_date)
    productions = _filter_rows_for_period(DatabaseHelper.list_productions(), "DateProduction", start_date, end_date)
    payrolls = DatabaseHelper.list_payrolls(start_date=start_date, end_date=end_date)
    payroll_summary = DatabaseHelper.get_workers_payroll_summary(start_date=start_date, end_date=end_date)
    production_summary = _summarize_production_rows(productions)

    expense_items_by_day: list[tuple[str, str]] = []
    paid_debts_items_by_day: list[tuple[str, str, str]] = []
    for row in cash_days:
        row_date = _parse_row_date(row.get("DateCaisse"))
        row_date_label = _format_date(row_date) if row_date is not None else _safe_text(row.get("DateCaisse"))
        for item in split_structured_lines(_safe_text(row.get("DepensesEffectuees")).strip()):
            expense_items_by_day.append((row_date_label, item))
        for name, amount in parse_named_amount_lines(_safe_text(row.get("DettesPayeesDetails")).strip()):
            paid_debts_items_by_day.append((row_date_label, name, amount or "-"))

    total_trays = sum(int(row.get("NombreBacs", 0) or 0) for row in orders)
    total_expected = sum(float(row.get("MontantAPercevoir", 0) or 0) for row in orders)
    total_received = sum(_order_accounted_received(row) for row in orders)
    total_received_gross = sum(_order_gross_received(row) for row in orders)
    total_debts = sum(float(row.get("Dette", 0) or 0) for row in orders)
    total_expenses = sum(float(row.get("MontantTotalDepenses", 0) or 0) for row in cash_days)
    paid_debts_period = sum(float(row.get("DettesPayeesAujourdHui", 0) or 0) for row in cash_days)
    total_entries = total_received + paid_debts_period
    balance = total_entries - total_expenses
    total_commissions = sum(float(row.get("Commissions", 0) or 0) for row in commissions)
    total_net_commissions = sum(float(row.get("NetAPayer", 0) or 0) for row in commissions)
    total_payroll_net = _payroll_total(payrolls, "MontantNet")
    balance_after_commissions = balance - total_net_commissions
    remaining_accumulated_debts = float(cash_days[-1].get("DettesAccumuleesRestantes", 0) or 0) if cash_days else 0.0

    return {
        "start_date": start_date,
        "end_date": end_date,
        "period_label": f"Du {_format_date(start_date)} au {_format_date(end_date)}",
        "report_title": f"RAPPORT DE PERIODE - DU {_format_date(start_date)} AU {_format_date(end_date)}",
        "role": normalized_role,
        "allowed_sections": allowed_sections,
        "scope_label": scope_label,
        "scope_description": scope_description,
        "generated_by": generated_by.strip() or "Utilisateur non identifié",
        "generated_role": (generated_role or normalized_role).strip() or normalized_role,
        "stock_journals": stock_journals,
        "stock_exits": stock_exits,
        "stock_supplies": stock_supplies,
        "orders": orders,
        "cash_days": cash_days,
        "commissions": commissions,
        "payrolls": payrolls,
        "payroll_summary": payroll_summary,
        "productions": productions,
        "production_summary": production_summary,
        "expense_items_by_day": expense_items_by_day,
        "paid_debts_items_by_day": paid_debts_items_by_day,
        "total_expected": total_expected,
        "total_received": total_received,
        "total_received_gross": total_received_gross,
        "total_debts": total_debts,
        "total_trays": total_trays,
        "total_expenses": total_expenses,
        "paid_debts_period": paid_debts_period,
        "total_entries": total_entries,
        "balance": balance,
        "remaining_accumulated_debts": remaining_accumulated_debts,
        "total_commissions": total_commissions,
        "total_net_commissions": total_net_commissions,
        "balance_after_commissions": balance_after_commissions,
        "total_payroll_net": total_payroll_net,
        "balance_after_payrolls": balance - total_payroll_net,
        "total_farine": sum(float(row.get("SacsUtilises", 0) or 0) for row in stock_exits),
        "total_levure": sum(float(row.get("PaquetsUtilises", 0) or 0) for row in stock_exits),
        "total_sel": sum(float(row.get("KgSelUtilises", 0) or 0) for row in stock_exits),
        "total_huile": sum(float(row.get("LitresHuileUtilises", 0) or 0) for row in stock_exits),
        "total_farine_added": sum(float(row.get("SacsAjoutes", 0) or 0) for row in stock_supplies),
        "total_levure_added": sum(float(row.get("PaquetsAjoutes", 0) or 0) for row in stock_supplies),
        "total_sel_added": sum(float(row.get("KgSelAjoutes", 0) or 0) for row in stock_supplies),
        "total_huile_added": sum(float(row.get("LitresHuileAjoutes", 0) or 0) for row in stock_supplies),
    }


def _build_monthly_summary_sheet(workbook: Workbook, context: dict[str, Any]) -> None:
    sheet = workbook.active
    sheet.title = "Résumé"
    sheet.freeze_panes = "A13"
    _add_sheet_watermark(sheet, "D7", 260, 260)
    start_row = _apply_brand_header(
        sheet,
        context["period_label"],
        context["report_title"],
        context["scope_label"],
        context["scope_description"],
        context["generated_by"],
        context["generated_role"],
    )
    sheet.cell(start_row, 1, "Indicateur")
    sheet.cell(start_row, 2, "Valeur")
    sheet.cell(start_row, 3, "Type")
    for col in range(1, 4):
        _apply_cell_style(sheet.cell(start_row, col), fill=HEADER_FILL, alignment=Alignment(horizontal="left"), font=HEADER_FONT)

    rows: list[tuple[str, int | float, str]] = []
    if "stock" in context["allowed_sections"]:
        rows.extend(
            [
                ("Jours avec journal de stock", len(context["stock_journals"]), "nombre"),
                ("Approvisionnements du mois", len(context["stock_supplies"]), "nombre"),
                ("Sorties de stock du mois", len(context["stock_exits"]), "nombre"),
            ]
        )
    if "production" in context["allowed_sections"]:
        production = context["production_summary"]
        rows.extend(
            [
                ("Jours de production saisis", len(context["productions"]), "nombre"),
                ("Bacs commandés", production["NombreBacsCommandes"], "nombre"),
                ("Total bacs produits", production["NombreBacsProduits"], "nombre"),
                ("Nombre de sacs utilisés", production["NombreSacsUtilises"], "nombre"),
                ("Écart production", production["EcartCommandes"], "nombre"),
            ]
        )
    if "orders" in context["allowed_sections"]:
        rows.extend(
            [
                ("Commandes du mois", len(context["orders"]), "nombre"),
                ("Total bacs", context["total_trays"], "nombre"),
                ("Montant attendu", context["total_expected"], "monnaie"),
                ("Payé par clients", context["total_received_gross"], "monnaie"),
                ("Reçu commandes", context["total_received"], "monnaie"),
                ("Dettes", context["total_debts"], "monnaie"),
            ]
        )
    if "cash" in context["allowed_sections"]:
        rows.extend(
            [
                ("Dettes payées du mois", context["paid_debts_month"], "monnaie"),
                ("Dettes accumulées restantes", context["remaining_accumulated_debts"], "monnaie"),
                ("Total des entrées", context["total_entries"], "monnaie"),
                ("Dépenses du mois", context["total_expenses"], "monnaie"),
                ("Solde du mois", context["balance"], "monnaie"),
            ]
        )
    if "commissions" in context["allowed_sections"]:
        rows.extend(
            [
                ("Commissions", context["total_commissions"], "monnaie"),
                ("Net à payer des commissions", context["total_net_commissions"], "monnaie"),
                *(
                    [("Solde après paiement des commissions", context["balance_after_commissions"], "monnaie")]
                    if "cash" in context["allowed_sections"]
                    else []
                ),
            ]
        )
    if "workers" in context["allowed_sections"]:
        payroll_summary = context["payroll_summary"]
        rows.extend(
            [
                ("Travailleurs actifs", int(payroll_summary.get("TravailleursActifs", 0) or 0), "nombre"),
                ("Masse salariale mensuelle", float(payroll_summary.get("MasseSalarialeMensuelle", 0) or 0), "monnaie"),
                ("Paies travailleurs", context["total_payroll_net"], "monnaie"),
                ("Solde après paies", context["balance_after_payrolls"], "monnaie"),
            ]
        )

    for row_offset, (label, value, kind) in enumerate(rows, start=1):
        row_index = start_row + row_offset
        sheet.cell(row_index, 1, label)
        value_cell = sheet.cell(row_index, 2, value)
        sheet.cell(row_index, 3, kind)
        _apply_cell_style(sheet.cell(row_index, 1), alignment=Alignment(horizontal="left"))
        _apply_cell_style(sheet.cell(row_index, 2), alignment=Alignment(horizontal="left"))
        _apply_cell_style(sheet.cell(row_index, 3), alignment=Alignment(horizontal="left"))
        if kind == "monnaie":
            value_cell.number_format = MONEY_FORMAT
        _apply_cash_emphasis(sheet, row_index, label, end_column=2)

    end_row = start_row + len(rows)
    _add_table(sheet, start_row, end_row, 3, "ResumeMensuel")
    sheet.column_dimensions["C"].hidden = True
    _autofit_columns(sheet, min_width=14, max_width=34)


def _build_monthly_stock_sheet(workbook: Workbook, context: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("Stock")
    sheet.freeze_panes = "A6"
    _add_sheet_watermark(sheet, "G6", 220, 220)

    sheet["A1"] = "Stock du mois"
    _apply_cell_style(sheet["A1"], fill=TITLE_FILL, alignment=Alignment(horizontal="left"), font=TITLE_FONT)
    sheet["A2"] = f"Période : {context['period_label']}"
    _apply_cell_style(sheet["A2"], alignment=Alignment(horizontal="left"))

    summary_headers = ["Indicateur", "Valeur"]
    for col_index, header in enumerate(summary_headers, start=1):
        _apply_cell_style(sheet.cell(4, col_index, header), fill=HEADER_FILL, alignment=Alignment(horizontal="left"), font=HEADER_FONT)
    summary_rows = [
        ("Jours avec journal", len(context["stock_journals"])),
        ("Approvisionnements enregistrés", len(context["stock_supplies"])),
        ("Sorties enregistrées", len(context["stock_exits"])),
        ("Farine ajoutée", context["total_farine_added"]),
        ("Levure ajoutée", context["total_levure_added"]),
        ("Sel ajouté", context["total_sel_added"]),
        ("Huile ajoutée", context["total_huile_added"]),
        ("Farine utilisée", context["total_farine"]),
        ("Levure utilisée", context["total_levure"]),
        ("Sel utilisé", context["total_sel"]),
        ("Huile utilisée", context["total_huile"]),
    ]
    current_row = 4
    for label, value in summary_rows:
        current_row += 1
        _apply_cell_style(sheet.cell(current_row, 1, label), alignment=Alignment(horizontal="left"))
        _apply_cell_style(sheet.cell(current_row, 2, value), alignment=Alignment(horizontal="left"))
    _add_table(sheet, 4, current_row, 2, "StockMensuel")

    current_row += 3
    sheet.cell(current_row, 1, "Approvisionnements")
    _apply_cell_style(sheet.cell(current_row, 1), fill=SECTION_FILL, alignment=Alignment(horizontal="left"), font=SECTION_FONT)
    current_row += 1
    headers = ["Date", "Farine", "Levure", "Sel", "Huile"]
    for col_index, header in enumerate(headers, start=1):
        _apply_cell_style(sheet.cell(current_row, col_index, header), fill=HEADER_FILL, alignment=Alignment(horizontal="left"), font=HEADER_FONT)
    start_table = current_row
    if context["stock_supplies"]:
        for row in context["stock_supplies"]:
            current_row += 1
            row_date = _parse_row_date(row.get("DateApprovisionnement"))
            values = [
                _format_date(row_date) if row_date is not None else _safe_text(row.get("DateApprovisionnement")),
                float(row.get("SacsAjoutes", 0) or 0),
                float(row.get("PaquetsAjoutes", 0) or 0),
                float(row.get("KgSelAjoutes", 0) or 0),
                float(row.get("LitresHuileAjoutes", 0) or 0),
            ]
            for col_index, value in enumerate(values, start=1):
                _apply_cell_style(sheet.cell(current_row, col_index, value), alignment=Alignment(horizontal="left"))
        _add_table(sheet, start_table, current_row, 5, "ApprovisionnementsStockMensuel")
    else:
        current_row += 1
        _apply_cell_style(sheet.cell(current_row, 1, "Aucun approvisionnement enregistré pour ce mois."), alignment=Alignment(horizontal="left"), font=NOTE_FONT)

    current_row += 3
    sheet.cell(current_row, 1, "Sorties de stock")
    _apply_cell_style(sheet.cell(current_row, 1), fill=SECTION_FILL, alignment=Alignment(horizontal="left"), font=SECTION_FONT)
    current_row += 1
    headers = ["Date", "Farine", "Levure", "Sel", "Huile"]
    for col_index, header in enumerate(headers, start=1):
        _apply_cell_style(sheet.cell(current_row, col_index, header), fill=HEADER_FILL, alignment=Alignment(horizontal="left"), font=HEADER_FONT)
    start_table = current_row
    if context["stock_exits"]:
        for row in context["stock_exits"]:
            current_row += 1
            row_date = _parse_row_date(row.get("DateSortie"))
            values = [
                _format_date(row_date) if row_date is not None else _safe_text(row.get("DateSortie")),
                float(row.get("SacsUtilises", 0) or 0),
                float(row.get("PaquetsUtilises", 0) or 0),
                float(row.get("KgSelUtilises", 0) or 0),
                float(row.get("LitresHuileUtilises", 0) or 0),
            ]
            for col_index, value in enumerate(values, start=1):
                _apply_cell_style(sheet.cell(current_row, col_index, value), alignment=Alignment(horizontal="left"))
        _add_table(sheet, start_table, current_row, 5, "SortiesStockMensuel")
    else:
        current_row += 1
        _apply_cell_style(sheet.cell(current_row, 1, "Aucune sortie de stock enregistrée pour ce mois."), alignment=Alignment(horizontal="left"), font=NOTE_FONT)

    _autofit_columns(sheet)


def _build_production_period_sheet(workbook: Workbook, context: dict[str, Any], title: str, table_name: str) -> None:
    sheet = workbook.create_sheet("Production")
    sheet.freeze_panes = "A4"
    _add_sheet_watermark(sheet, "G5", 220, 220)

    sheet["A1"] = title
    _apply_cell_style(sheet["A1"], fill=TITLE_FILL, alignment=Alignment(horizontal="left"), font=TITLE_FONT)
    sheet["A2"] = f"Période : {context['period_label']}"
    _apply_cell_style(sheet["A2"], alignment=Alignment(horizontal="left"))

    summary_headers = ["Champ", "Valeur"]
    for col_index, header in enumerate(summary_headers, start=1):
        _apply_cell_style(sheet.cell(4, col_index, header), fill=HEADER_FILL, alignment=Alignment(horizontal="left"), font=HEADER_FONT)
    current_row = 4
    for label, value in _production_field_rows(context["production_summary"]):
        current_row += 1
        _apply_cell_style(sheet.cell(current_row, 1, label), alignment=Alignment(horizontal="left"))
        _apply_cell_style(sheet.cell(current_row, 2, value), alignment=Alignment(horizontal="left", wrap_text=True))
    _add_table(sheet, 4, current_row, 2, f"{table_name}Synthese")

    current_row += 3
    sheet.cell(current_row, 1, "Détail par jour")
    _apply_cell_style(sheet.cell(current_row, 1), fill=SECTION_FILL, alignment=Alignment(horizontal="left"), font=SECTION_FONT)
    current_row += 1
    headers = [
        "Date",
        "Commandés",
        "Livrés dép.",
        "Livrés mamans",
        "Donnés",
        "Éch.",
        "Restants",
        "Foutus",
        "Produits",
        "Sacs",
        "Écart",
    ]
    for col_index, header in enumerate(headers, start=1):
        _apply_cell_style(sheet.cell(current_row, col_index, header), fill=HEADER_FILL, alignment=Alignment(horizontal="left"), font=HEADER_FONT)
    start_table = current_row
    if context["productions"]:
        for row in context["productions"]:
            current_row += 1
            row_date = _parse_row_date(row.get("DateProduction"))
            values = [
                _format_date(row_date) if row_date is not None else _safe_text(row.get("DateProduction")),
                float(row.get("NombreBacsCommandes", 0) or 0),
                float(row.get("NombreBacsLivresDepositaires", 0) or 0),
                float(row.get("NombreBacsLivresMamans", 0) or 0),
                float(row.get("NombreBacsDonnes", 0) or 0),
                float(row.get("NombreEchantillons", 0) or 0),
                float(row.get("NombreBacsRestants", 0) or 0),
                float(row.get("NombreBacsFoutus", 0) or 0),
                float(row.get("NombreBacsProduits", 0) or 0),
                float(row.get("NombreSacsUtilises", 0) or 0),
                float(row.get("EcartCommandes", 0) or 0),
            ]
            for col_index, value in enumerate(values, start=1):
                _apply_cell_style(sheet.cell(current_row, col_index, value), alignment=Alignment(horizontal="left"))
        _add_table(sheet, start_table, current_row, 11, table_name)
    else:
        current_row += 1
        _apply_cell_style(sheet.cell(current_row, 1, "Aucune production enregistrée pour cette période."), alignment=Alignment(horizontal="left"), font=NOTE_FONT)

    _autofit_columns(sheet)


def _build_monthly_orders_sheet(workbook: Workbook, context: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("Commandes")
    sheet.freeze_panes = "A4"
    _add_sheet_watermark(sheet, "G5", 220, 220)

    sheet["A1"] = "Synthèse des commandes du mois"
    _apply_cell_style(sheet["A1"], fill=TITLE_FILL, alignment=Alignment(horizontal="left"), font=TITLE_FONT)
    sheet["A2"] = f"Période : {context['period_label']}"
    _apply_cell_style(sheet["A2"], alignment=Alignment(horizontal="left"))

    headers = ["Statut", "Commandes", "Bacs", "À percevoir", "Reçu commande", "Avance utilisée", "Nouvelle avance", "Dette"]
    for col_index, header in enumerate(headers, start=1):
        _apply_cell_style(sheet.cell(3, col_index, header), fill=HEADER_FILL, alignment=Alignment(horizontal="left"), font=HEADER_FONT)

    current_row = 3
    summary_rows = _order_status_summary_rows(context["orders"])
    if summary_rows:
        for status, count, trays, expected, received, advance_used, advance_generated, debt in summary_rows:
            current_row += 1
            values = [
                status,
                count,
                trays,
                expected,
                received,
                advance_used,
                advance_generated,
                debt,
            ]
            for col_index, value in enumerate(values, start=1):
                _apply_cell_style(sheet.cell(current_row, col_index, value), alignment=Alignment(horizontal="left"))
            for col in range(4, 9):
                sheet.cell(current_row, col).number_format = MONEY_FORMAT
            if status == "Total":
                for col_index in range(1, 9):
                    _apply_cell_style(
                        sheet.cell(current_row, col_index),
                        fill=SECTION_FILL,
                        alignment=Alignment(horizontal="left"),
                        font=SECTION_FONT if col_index == 1 else BODY_FONT,
                    )
                for col in range(4, 9):
                    sheet.cell(current_row, col).number_format = MONEY_FORMAT
        _add_table(sheet, 3, current_row, 8, "CommandesMensuellesSynthese")

        current_row += 2
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 1, end_column=8)
        note = (
            "La liste détaillée de toutes les commandes du mois n'est pas affichée ici afin de garder "
            "le rapport mensuel lisible. Le détail reste disponible dans les rapports journaliers ou de période."
        )
        _apply_cell_style(
            sheet.cell(current_row, 1, note),
            alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
            font=NOTE_FONT,
        )
    else:
        _apply_cell_style(sheet["A4"], alignment=Alignment(horizontal="left"), font=NOTE_FONT)
        sheet["A4"] = "Aucune commande enregistrée pour ce mois."

    _autofit_columns(sheet)


def _build_monthly_cash_sheet(workbook: Workbook, context: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("Caisse")
    sheet.freeze_panes = "A4"
    _add_sheet_watermark(sheet, "G5", 220, 220)

    sheet["A1"] = "Caisse du mois"
    _apply_cell_style(sheet["A1"], fill=TITLE_FILL, alignment=Alignment(horizontal="left"), font=TITLE_FONT)
    sheet["A2"] = f"Période : {context['period_label']}"
    _apply_cell_style(sheet["A2"], alignment=Alignment(horizontal="left"))

    headers = ["Champ", "Valeur"]
    for col_index, header in enumerate(headers, start=1):
        _apply_cell_style(sheet.cell(4, col_index, header), fill=HEADER_FILL, alignment=Alignment(horizontal="left"), font=HEADER_FONT)

    rows = [
        ("Montant attendu", context["total_expected"]),
        ("Payé par clients", context["total_received_gross"]),
        ("Reçu commandes", context["total_received"]),
        ("Dettes", context["total_debts"]),
        ("Dettes payées du mois", context["paid_debts_month"]),
        ("Dettes accumulées restantes", context["remaining_accumulated_debts"]),
        ("Total des entrées", context["total_entries"]),
        ("Dépenses du mois", context["total_expenses"]),
        *(
            [
                ("Paies travailleurs", context["total_payroll_net"]),
                ("Total des sorties", context["total_expenses"] + context["total_payroll_net"]),
                ("Solde après paies", context["balance_after_payrolls"]),
            ]
            if "workers" in context["allowed_sections"]
            else []
        ),
        ("Solde du mois", context["balance"]),
    ]
    current_row = 4
    for label, value in rows:
        current_row += 1
        _apply_cell_style(sheet.cell(current_row, 1, label), alignment=Alignment(horizontal="left"))
        _apply_cell_style(sheet.cell(current_row, 2, value), alignment=Alignment(horizontal="left"), number_format=MONEY_FORMAT)
        _apply_cash_emphasis(sheet, current_row, label, end_column=2)
    _add_table(sheet, 4, current_row, 2, "CaisseMensuelle")

    current_row += 3
    sheet.cell(current_row, 1, "Synthèse journalière")
    _apply_cell_style(sheet.cell(current_row, 1), fill=SECTION_FILL, alignment=Alignment(horizontal="left"), font=SECTION_FONT)
    current_row += 1
    day_headers = ["Date", "Reçu commandes", "Dettes payées", "Entrées", "Dépenses", "Solde", "Dettes restantes"]
    for col_index, header in enumerate(day_headers, start=1):
        _apply_cell_style(sheet.cell(current_row, col_index, header), fill=HEADER_FILL, alignment=Alignment(horizontal="left"), font=HEADER_FONT)
    start_daily = current_row
    if context["cash_days"]:
        for row in context["cash_days"]:
            current_row += 1
            row_date = _parse_row_date(row.get("DateCaisse"))
            values = [
                _format_date(row_date) if row_date is not None else _safe_text(row.get("DateCaisse")),
                float(row.get("MontantRecu", 0) or 0),
                float(row.get("DettesPayeesAujourdHui", 0) or 0),
                float(row.get("TotalEntrees", 0) or 0),
                float(row.get("MontantTotalDepenses", 0) or 0),
                float(row.get("Solde", 0) or 0),
                float(row.get("DettesAccumuleesRestantes", 0) or 0),
            ]
            for col_index, value in enumerate(values, start=1):
                _apply_cell_style(sheet.cell(current_row, col_index, value), alignment=Alignment(horizontal="left"))
            for col in range(2, 8):
                sheet.cell(current_row, col).number_format = MONEY_FORMAT
        _add_table(sheet, start_daily, current_row, 7, "CaisseJoursMensuels")
    else:
        current_row += 1
        _apply_cell_style(sheet.cell(current_row, 1, "Aucune fiche de caisse enregistrée pour ce mois."), alignment=Alignment(horizontal="left"), font=NOTE_FONT)

    current_row += 3
    sheet.cell(current_row, 1, "Liste mensuelle des dépenses")
    _apply_cell_style(sheet.cell(current_row, 1), fill=SECTION_FILL, alignment=Alignment(horizontal="left"), font=SECTION_FONT)
    current_row += 1
    headers = ["Date", "Détail"]
    for col_index, header in enumerate(headers, start=1):
        _apply_cell_style(sheet.cell(current_row, col_index, header), fill=HEADER_FILL, alignment=Alignment(horizontal="left"), font=HEADER_FONT)
    start_expenses = current_row
    if context["expense_items_by_day"]:
        for row_date, detail in context["expense_items_by_day"]:
            current_row += 1
            _apply_cell_style(sheet.cell(current_row, 1, row_date), alignment=Alignment(horizontal="left"))
            _apply_cell_style(sheet.cell(current_row, 2, detail), alignment=Alignment(horizontal="left", wrap_text=True))
        _add_table(sheet, start_expenses, current_row, 2, "DepensesMensuelles")
    else:
        current_row += 1
        _apply_cell_style(sheet.cell(current_row, 1, "Aucune dépense détaillée enregistrée pour ce mois."), alignment=Alignment(horizontal="left"), font=NOTE_FONT)

    current_row += 3
    sheet.cell(current_row, 1, "NB")
    _apply_cell_style(sheet.cell(current_row, 1), fill=SECTION_FILL, alignment=Alignment(horizontal="left"), font=SECTION_FONT)
    current_row += 1
    payroll_sentence = ""
    if "workers" in context["allowed_sections"]:
        payroll_sentence = (
            f" Les paies des travailleurs représentent {_format_fc(context['total_payroll_net'])}; "
            f"le solde réel après paies est de {_format_fc(context['balance_after_payrolls'])}."
        )
    recap_text = (
        f"Pour le mois {context['month_label']}, les entrées correspondent au reçu commandes ({_format_fc(context['total_received'])}) "
        f"additionné aux dettes payées ({_format_fc(context['paid_debts_month'])}), soit un total des entrées de {_format_fc(context['total_entries'])}. "
        f"Les sorties correspondent aux dépenses du mois, soit {_format_fc(context['total_expenses'])}. "
        f"Le solde mensuel ressort donc à {_format_fc(context['balance'])}. "
        f"Les dettes accumulées restantes à la fin du mois sont de {_format_fc(context['remaining_accumulated_debts'])}."
        f"{payroll_sentence}"
    )
    sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 2, end_column=6)
    _apply_cell_style(sheet.cell(current_row, 1, recap_text), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True))

    _autofit_columns(sheet)


def _build_monthly_commissions_sheet(workbook: Workbook, context: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("Commissions")
    sheet.freeze_panes = "A4"
    _add_sheet_watermark(sheet, "G5", 220, 220)

    sheet["A1"] = "Commissions du mois"
    _apply_cell_style(sheet["A1"], fill=TITLE_FILL, alignment=Alignment(horizontal="left"), font=TITLE_FONT)
    sheet["A2"] = f"Période : {context['period_label']}"
    _apply_cell_style(sheet["A2"], alignment=Alignment(horizontal="left"))

    headers = ["Date", "Nom", "Statut", "Bacs", "Payé", "Commission", "Dette", "Net"]
    for col_index, header in enumerate(headers, start=1):
        _apply_cell_style(sheet.cell(3, col_index, header), fill=HEADER_FILL, alignment=Alignment(horizontal="left"), font=HEADER_FONT)

    current_row = 3
    if context["commissions"]:
        for item in context["commissions"]:
            current_row += 1
            row_date = _parse_row_date(item.get("DateCommission"))
            values = [
                _format_date(row_date) if row_date is not None else _safe_text(item.get("DateCommission")),
                _safe_text(item.get("Nom")),
                normalize_status_form_label(item.get("Statut")),
                int(item.get("NombreBacs", 0) or 0),
                float(item.get("MontantPaye", 0) or 0),
                float(item.get("Commissions", 0) or 0),
                float(item.get("Dettes", 0) or 0),
                float(item.get("NetAPayer", 0) or 0),
            ]
            for col_index, value in enumerate(values, start=1):
                _apply_cell_style(sheet.cell(current_row, col_index, value), alignment=Alignment(horizontal="left"))
            for col in range(5, 9):
                sheet.cell(current_row, col).number_format = MONEY_FORMAT

        totals_row = current_row + 1
        sheet.cell(totals_row, 1, "Totaux")
        sheet.cell(totals_row, 4, f"=SUM(D4:D{current_row})")
        sheet.cell(totals_row, 5, f"=SUM(E4:E{current_row})")
        sheet.cell(totals_row, 6, f"=SUM(F4:F{current_row})")
        sheet.cell(totals_row, 7, f"=SUM(G4:G{current_row})")
        sheet.cell(totals_row, 8, f"=SUM(H4:H{current_row})")
        for col_index in range(1, 9):
            _apply_cell_style(sheet.cell(totals_row, col_index), fill=SECTION_FILL, alignment=Alignment(horizontal="left"), font=SECTION_FONT if col_index == 1 else BODY_FONT)
        for col in range(5, 9):
            sheet.cell(totals_row, col).number_format = MONEY_FORMAT
        _add_table(sheet, 3, current_row, 8, "CommissionsMensuelles")
    else:
        _apply_cell_style(sheet["A4"], alignment=Alignment(horizontal="left"), font=NOTE_FONT)
        sheet["A4"] = "Aucune commission enregistrée pour ce mois."

    _autofit_columns(sheet)


def _build_period_summary_sheet(workbook: Workbook, context: dict[str, Any]) -> None:
    sheet = workbook.active
    sheet.title = "Résumé"
    sheet.freeze_panes = "A13"
    _add_sheet_watermark(sheet, "D7", 260, 260)
    start_row = _apply_brand_header(
        sheet,
        context["period_label"],
        context["report_title"],
        context["scope_label"],
        context["scope_description"],
        context["generated_by"],
        context["generated_role"],
    )
    sheet.cell(start_row, 1, "Indicateur")
    sheet.cell(start_row, 2, "Valeur")
    sheet.cell(start_row, 3, "Type")
    for col in range(1, 4):
        _apply_cell_style(sheet.cell(start_row, col), fill=HEADER_FILL, alignment=Alignment(horizontal="left"), font=HEADER_FONT)

    rows: list[tuple[str, int | float, str]] = []
    if "stock" in context["allowed_sections"]:
        rows.extend(
            [
                ("Jours avec journal de stock", len(context["stock_journals"]), "nombre"),
                ("Approvisionnements sur la période", len(context["stock_supplies"]), "nombre"),
                ("Sorties de stock sur la période", len(context["stock_exits"]), "nombre"),
            ]
        )
    if "production" in context["allowed_sections"]:
        production = context["production_summary"]
        rows.extend(
            [
                ("Jours de production saisis", len(context["productions"]), "nombre"),
                ("Bacs commandés", production["NombreBacsCommandes"], "nombre"),
                ("Total bacs produits", production["NombreBacsProduits"], "nombre"),
                ("Nombre de sacs utilisés", production["NombreSacsUtilises"], "nombre"),
                ("Écart production", production["EcartCommandes"], "nombre"),
            ]
        )
    if "orders" in context["allowed_sections"]:
        rows.extend(
            [
                ("Commandes sur la période", len(context["orders"]), "nombre"),
                ("Total bacs", context["total_trays"], "nombre"),
                ("Montant attendu", context["total_expected"], "monnaie"),
                ("Payé par clients", context["total_received_gross"], "monnaie"),
                ("Reçu commandes", context["total_received"], "monnaie"),
                ("Dettes", context["total_debts"], "monnaie"),
            ]
        )
    if "cash" in context["allowed_sections"]:
        rows.extend(
            [
                ("Dettes payées sur la période", context["paid_debts_period"], "monnaie"),
                ("Dettes accumulées restantes", context["remaining_accumulated_debts"], "monnaie"),
                ("Total des entrées", context["total_entries"], "monnaie"),
                ("Dépenses sur la période", context["total_expenses"], "monnaie"),
                ("Solde sur la période", context["balance"], "monnaie"),
            ]
        )
    if "commissions" in context["allowed_sections"]:
        rows.extend(
            [
                ("Commissions", context["total_commissions"], "monnaie"),
                ("Net à payer des commissions", context["total_net_commissions"], "monnaie"),
                *(
                    [("Solde après paiement des commissions", context["balance_after_commissions"], "monnaie")]
                    if "cash" in context["allowed_sections"]
                    else []
                ),
            ]
        )
    if "workers" in context["allowed_sections"]:
        payroll_summary = context["payroll_summary"]
        rows.extend(
            [
                ("Travailleurs actifs", int(payroll_summary.get("TravailleursActifs", 0) or 0), "nombre"),
                ("Masse salariale mensuelle", float(payroll_summary.get("MasseSalarialeMensuelle", 0) or 0), "monnaie"),
                ("Paies travailleurs", context["total_payroll_net"], "monnaie"),
                ("Solde après paies", context["balance_after_payrolls"], "monnaie"),
            ]
        )

    for row_offset, (label, value, kind) in enumerate(rows, start=1):
        row_index = start_row + row_offset
        sheet.cell(row_index, 1, label)
        value_cell = sheet.cell(row_index, 2, value)
        sheet.cell(row_index, 3, kind)
        _apply_cell_style(sheet.cell(row_index, 1), alignment=Alignment(horizontal="left"))
        _apply_cell_style(sheet.cell(row_index, 2), alignment=Alignment(horizontal="left"))
        _apply_cell_style(sheet.cell(row_index, 3), alignment=Alignment(horizontal="left"))
        if kind == "monnaie":
            value_cell.number_format = MONEY_FORMAT
        _apply_cash_emphasis(sheet, row_index, label, end_column=2)

    end_row = start_row + len(rows)
    _add_table(sheet, start_row, end_row, 3, "ResumePeriode")
    sheet.column_dimensions["C"].hidden = True
    _autofit_columns(sheet, min_width=14, max_width=34)


def _build_period_stock_sheet(workbook: Workbook, context: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("Stock")
    sheet.freeze_panes = "A6"
    _add_sheet_watermark(sheet, "G6", 220, 220)

    sheet["A1"] = "Stock sur la période"
    _apply_cell_style(sheet["A1"], fill=TITLE_FILL, alignment=Alignment(horizontal="left"), font=TITLE_FONT)
    sheet["A2"] = f"Période : {context['period_label']}"
    _apply_cell_style(sheet["A2"], alignment=Alignment(horizontal="left"))

    summary_headers = ["Indicateur", "Valeur"]
    for col_index, header in enumerate(summary_headers, start=1):
        _apply_cell_style(sheet.cell(4, col_index, header), fill=HEADER_FILL, alignment=Alignment(horizontal="left"), font=HEADER_FONT)
    summary_rows = [
        ("Jours avec journal", len(context["stock_journals"])),
        ("Approvisionnements enregistrés", len(context["stock_supplies"])),
        ("Sorties enregistrées", len(context["stock_exits"])),
        ("Farine ajoutée", context["total_farine_added"]),
        ("Levure ajoutée", context["total_levure_added"]),
        ("Sel ajouté", context["total_sel_added"]),
        ("Huile ajoutée", context["total_huile_added"]),
        ("Farine utilisée", context["total_farine"]),
        ("Levure utilisée", context["total_levure"]),
        ("Sel utilisé", context["total_sel"]),
        ("Huile utilisée", context["total_huile"]),
    ]
    current_row = 4
    for label, value in summary_rows:
        current_row += 1
        _apply_cell_style(sheet.cell(current_row, 1, label), alignment=Alignment(horizontal="left"))
        _apply_cell_style(sheet.cell(current_row, 2, value), alignment=Alignment(horizontal="left"))
    _add_table(sheet, 4, current_row, 2, "StockPeriode")

    current_row += 3
    sheet.cell(current_row, 1, "Approvisionnements")
    _apply_cell_style(sheet.cell(current_row, 1), fill=SECTION_FILL, alignment=Alignment(horizontal="left"), font=SECTION_FONT)
    current_row += 1
    headers = ["Date", "Farine", "Levure", "Sel", "Huile"]
    for col_index, header in enumerate(headers, start=1):
        _apply_cell_style(sheet.cell(current_row, col_index, header), fill=HEADER_FILL, alignment=Alignment(horizontal="left"), font=HEADER_FONT)
    start_table = current_row
    if context["stock_supplies"]:
        for row in context["stock_supplies"]:
            current_row += 1
            row_date = _parse_row_date(row.get("DateApprovisionnement"))
            values = [
                _format_date(row_date) if row_date is not None else _safe_text(row.get("DateApprovisionnement")),
                float(row.get("SacsAjoutes", 0) or 0),
                float(row.get("PaquetsAjoutes", 0) or 0),
                float(row.get("KgSelAjoutes", 0) or 0),
                float(row.get("LitresHuileAjoutes", 0) or 0),
            ]
            for col_index, value in enumerate(values, start=1):
                _apply_cell_style(sheet.cell(current_row, col_index, value), alignment=Alignment(horizontal="left"))
        _add_table(sheet, start_table, current_row, 5, "ApprovisionnementsStockPeriode")
    else:
        current_row += 1
        _apply_cell_style(sheet.cell(current_row, 1, "Aucun approvisionnement enregistré pour cette période."), alignment=Alignment(horizontal="left"), font=NOTE_FONT)

    current_row += 3
    sheet.cell(current_row, 1, "Sorties de stock")
    _apply_cell_style(sheet.cell(current_row, 1), fill=SECTION_FILL, alignment=Alignment(horizontal="left"), font=SECTION_FONT)
    current_row += 1
    headers = ["Date", "Farine", "Levure", "Sel", "Huile"]
    for col_index, header in enumerate(headers, start=1):
        _apply_cell_style(sheet.cell(current_row, col_index, header), fill=HEADER_FILL, alignment=Alignment(horizontal="left"), font=HEADER_FONT)
    start_table = current_row
    if context["stock_exits"]:
        for row in context["stock_exits"]:
            current_row += 1
            row_date = _parse_row_date(row.get("DateSortie"))
            values = [
                _format_date(row_date) if row_date is not None else _safe_text(row.get("DateSortie")),
                float(row.get("SacsUtilises", 0) or 0),
                float(row.get("PaquetsUtilises", 0) or 0),
                float(row.get("KgSelUtilises", 0) or 0),
                float(row.get("LitresHuileUtilises", 0) or 0),
            ]
            for col_index, value in enumerate(values, start=1):
                _apply_cell_style(sheet.cell(current_row, col_index, value), alignment=Alignment(horizontal="left"))
        _add_table(sheet, start_table, current_row, 5, "SortiesStockPeriode")
    else:
        current_row += 1
        _apply_cell_style(sheet.cell(current_row, 1, "Aucune sortie de stock enregistrée pour cette période."), alignment=Alignment(horizontal="left"), font=NOTE_FONT)

    _autofit_columns(sheet)


def _build_period_orders_sheet(workbook: Workbook, context: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("Commandes")
    sheet.freeze_panes = "A4"
    _add_sheet_watermark(sheet, "G5", 220, 220)

    sheet["A1"] = "Commandes sur la période"
    _apply_cell_style(sheet["A1"], fill=TITLE_FILL, alignment=Alignment(horizontal="left"), font=TITLE_FONT)
    sheet["A2"] = f"Période : {context['period_label']}"
    _apply_cell_style(sheet["A2"], alignment=Alignment(horizontal="left"))

    headers = ["Date", "Client", "Statut", "Bacs", "À percevoir", "Payé client", "Reçu commande", "Avance utilisée", "Nouvelle avance", "Solde avance", "Dette"]
    for col_index, header in enumerate(headers, start=1):
        _apply_cell_style(sheet.cell(3, col_index, header), fill=HEADER_FILL, alignment=Alignment(horizontal="left"), font=HEADER_FONT)

    current_row = 3
    if context["orders"]:
        for item in context["orders"]:
            current_row += 1
            row_date = _parse_row_date(item.get("DateCommande"))
            values = [
                _format_date(row_date) if row_date is not None else _safe_text(item.get("DateCommande")),
                _safe_text(item.get("Client")),
                normalize_status_form_label(item.get("Statut")),
                int(item.get("NombreBacs", 0) or 0),
                float(item.get("MontantAPercevoir", 0) or 0),
                _order_gross_received(item),
                _order_accounted_received(item),
                float(item.get("AvanceUtilisee", 0) or 0),
                float(item.get("AvanceGeneree", 0) or 0),
                float(item.get("SoldeAvance", 0) or 0),
                float(item.get("Dette", 0) or 0),
            ]
            for col_index, value in enumerate(values, start=1):
                _apply_cell_style(sheet.cell(current_row, col_index, value), alignment=Alignment(horizontal="left"))
            for col in range(5, 12):
                sheet.cell(current_row, col).number_format = MONEY_FORMAT

        totals_row = current_row + 1
        sheet.cell(totals_row, 1, "Totaux")
        sheet.cell(totals_row, 4, f"=SUM(D4:D{current_row})")
        sheet.cell(totals_row, 5, f"=SUM(E4:E{current_row})")
        sheet.cell(totals_row, 6, f"=SUM(F4:F{current_row})")
        sheet.cell(totals_row, 7, f"=SUM(G4:G{current_row})")
        sheet.cell(totals_row, 8, f"=SUM(H4:H{current_row})")
        sheet.cell(totals_row, 9, f"=SUM(I4:I{current_row})")
        sheet.cell(totals_row, 11, f"=SUM(K4:K{current_row})")
        for col_index in range(1, 12):
            _apply_cell_style(sheet.cell(totals_row, col_index), fill=SECTION_FILL, alignment=Alignment(horizontal="left"), font=SECTION_FONT if col_index == 1 else BODY_FONT)
        for col in range(5, 12):
            sheet.cell(totals_row, col).number_format = MONEY_FORMAT
        _add_table(sheet, 3, current_row, 11, "CommandesPeriode")
    else:
        _apply_cell_style(sheet["A4"], alignment=Alignment(horizontal="left"), font=NOTE_FONT)
        sheet["A4"] = "Aucune commande enregistrée pour cette période."

    _autofit_columns(sheet)


def _build_period_cash_sheet(workbook: Workbook, context: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("Caisse")
    sheet.freeze_panes = "A4"
    _add_sheet_watermark(sheet, "G5", 220, 220)

    sheet["A1"] = "Caisse sur la période"
    _apply_cell_style(sheet["A1"], fill=TITLE_FILL, alignment=Alignment(horizontal="left"), font=TITLE_FONT)
    sheet["A2"] = f"Période : {context['period_label']}"
    _apply_cell_style(sheet["A2"], alignment=Alignment(horizontal="left"))

    headers = ["Champ", "Valeur"]
    for col_index, header in enumerate(headers, start=1):
        _apply_cell_style(sheet.cell(4, col_index, header), fill=HEADER_FILL, alignment=Alignment(horizontal="left"), font=HEADER_FONT)

    rows = [
        ("Montant attendu", context["total_expected"]),
        ("Payé par clients", context["total_received_gross"]),
        ("Reçu commandes", context["total_received"]),
        ("Dettes", context["total_debts"]),
        ("Dettes payées sur la période", context["paid_debts_period"]),
        ("Dettes accumulées restantes", context["remaining_accumulated_debts"]),
        ("Total des entrées", context["total_entries"]),
        ("Dépenses sur la période", context["total_expenses"]),
        *(
            [
                ("Paies travailleurs", context["total_payroll_net"]),
                ("Total des sorties", context["total_expenses"] + context["total_payroll_net"]),
                ("Solde après paies", context["balance_after_payrolls"]),
            ]
            if "workers" in context["allowed_sections"]
            else []
        ),
        ("Solde sur la période", context["balance"]),
    ]
    current_row = 4
    for label, value in rows:
        current_row += 1
        _apply_cell_style(sheet.cell(current_row, 1, label), alignment=Alignment(horizontal="left"))
        _apply_cell_style(sheet.cell(current_row, 2, value), alignment=Alignment(horizontal="left"), number_format=MONEY_FORMAT)
        _apply_cash_emphasis(sheet, current_row, label, end_column=2)
    _add_table(sheet, 4, current_row, 2, "CaissePeriode")

    current_row += 3
    sheet.cell(current_row, 1, "Synthèse journalière")
    _apply_cell_style(sheet.cell(current_row, 1), fill=SECTION_FILL, alignment=Alignment(horizontal="left"), font=SECTION_FONT)
    current_row += 1
    day_headers = ["Date", "Reçu commandes", "Dettes payées", "Entrées", "Dépenses", "Solde", "Dettes restantes"]
    for col_index, header in enumerate(day_headers, start=1):
        _apply_cell_style(sheet.cell(current_row, col_index, header), fill=HEADER_FILL, alignment=Alignment(horizontal="left"), font=HEADER_FONT)
    start_daily = current_row
    if context["cash_days"]:
        for row in context["cash_days"]:
            current_row += 1
            row_date = _parse_row_date(row.get("DateCaisse"))
            values = [
                _format_date(row_date) if row_date is not None else _safe_text(row.get("DateCaisse")),
                float(row.get("MontantRecu", 0) or 0),
                float(row.get("DettesPayeesAujourdHui", 0) or 0),
                float(row.get("TotalEntrees", 0) or 0),
                float(row.get("MontantTotalDepenses", 0) or 0),
                float(row.get("Solde", 0) or 0),
                float(row.get("DettesAccumuleesRestantes", 0) or 0),
            ]
            for col_index, value in enumerate(values, start=1):
                _apply_cell_style(sheet.cell(current_row, col_index, value), alignment=Alignment(horizontal="left"))
            for col in range(2, 8):
                sheet.cell(current_row, col).number_format = MONEY_FORMAT
        _add_table(sheet, start_daily, current_row, 7, "CaisseJoursPeriode")
    else:
        current_row += 1
        _apply_cell_style(sheet.cell(current_row, 1, "Aucune fiche de caisse enregistrée pour cette période."), alignment=Alignment(horizontal="left"), font=NOTE_FONT)

    current_row += 3
    sheet.cell(current_row, 1, "Liste des dépenses sur la période")
    _apply_cell_style(sheet.cell(current_row, 1), fill=SECTION_FILL, alignment=Alignment(horizontal="left"), font=SECTION_FONT)
    current_row += 1
    headers = ["Date", "Détail"]
    for col_index, header in enumerate(headers, start=1):
        _apply_cell_style(sheet.cell(current_row, col_index, header), fill=HEADER_FILL, alignment=Alignment(horizontal="left"), font=HEADER_FONT)
    start_expenses = current_row
    if context["expense_items_by_day"]:
        for row_date, detail in context["expense_items_by_day"]:
            current_row += 1
            _apply_cell_style(sheet.cell(current_row, 1, row_date), alignment=Alignment(horizontal="left"))
            _apply_cell_style(sheet.cell(current_row, 2, detail), alignment=Alignment(horizontal="left", wrap_text=True))
        _add_table(sheet, start_expenses, current_row, 2, "DepensesPeriode")
    else:
        current_row += 1
        _apply_cell_style(sheet.cell(current_row, 1, "Aucune dépense détaillée enregistrée pour cette période."), alignment=Alignment(horizontal="left"), font=NOTE_FONT)

    current_row += 3
    sheet.cell(current_row, 1, "Ceux qui ont payé leurs dettes")
    _apply_cell_style(sheet.cell(current_row, 1), fill=SECTION_FILL, alignment=Alignment(horizontal="left"), font=SECTION_FONT)
    current_row += 1
    paid_headers = ["Date", "Nom", "Montant payé"]
    for col_index, header in enumerate(paid_headers, start=1):
        _apply_cell_style(sheet.cell(current_row, col_index, header), fill=HEADER_FILL, alignment=Alignment(horizontal="left"), font=HEADER_FONT)
    start_paid = current_row
    if context["paid_debts_items_by_day"]:
        for row_date, name, amount in context["paid_debts_items_by_day"]:
            current_row += 1
            _apply_cell_style(sheet.cell(current_row, 1, row_date), alignment=Alignment(horizontal="left"))
            _apply_cell_style(sheet.cell(current_row, 2, name), alignment=Alignment(horizontal="left", wrap_text=True))
            _apply_cell_style(sheet.cell(current_row, 3, amount), alignment=Alignment(horizontal="left"))
        _add_table(sheet, start_paid, current_row, 3, "DettesPayeesPeriode")
    else:
        current_row += 1
        _apply_cell_style(sheet.cell(current_row, 1, "Aucun paiement de dette détaillé pour cette période."), alignment=Alignment(horizontal="left"), font=NOTE_FONT)

    current_row += 3
    sheet.cell(current_row, 1, "NB")
    _apply_cell_style(sheet.cell(current_row, 1), fill=SECTION_FILL, alignment=Alignment(horizontal="left"), font=SECTION_FONT)
    current_row += 1
    payroll_sentence = ""
    if "workers" in context["allowed_sections"]:
        payroll_sentence = (
            f" Les paies des travailleurs représentent {_format_fc(context['total_payroll_net'])}; "
            f"le solde réel après paies est de {_format_fc(context['balance_after_payrolls'])}."
        )
    recap_text = (
        f"Pour la période du {_format_date(context['start_date'])} au {_format_date(context['end_date'])}, les entrées correspondent au montant reçu "
        f"({_format_fc(context['total_received'])}) additionné aux dettes payées ({_format_fc(context['paid_debts_period'])}), soit un total des entrées "
        f"de {_format_fc(context['total_entries'])}. Les sorties correspondent aux dépenses enregistrées sur la période, soit "
        f"{_format_fc(context['total_expenses'])}. Le solde ressort donc à {_format_fc(context['balance'])}. "
        f"Les dettes accumulées restantes à la fin de la période sont de {_format_fc(context['remaining_accumulated_debts'])}."
        f"{payroll_sentence}"
    )
    sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 2, end_column=6)
    _apply_cell_style(sheet.cell(current_row, 1, recap_text), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True))

    _autofit_columns(sheet)


def _build_period_commissions_sheet(workbook: Workbook, context: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("Commissions")
    sheet.freeze_panes = "A4"
    _add_sheet_watermark(sheet, "G5", 220, 220)

    sheet["A1"] = "Commissions sur la période"
    _apply_cell_style(sheet["A1"], fill=TITLE_FILL, alignment=Alignment(horizontal="left"), font=TITLE_FONT)
    sheet["A2"] = f"Période : {context['period_label']}"
    _apply_cell_style(sheet["A2"], alignment=Alignment(horizontal="left"))

    headers = ["Date", "Nom", "Statut", "Bacs", "Payé", "Commission", "Dette", "Net"]
    for col_index, header in enumerate(headers, start=1):
        _apply_cell_style(sheet.cell(3, col_index, header), fill=HEADER_FILL, alignment=Alignment(horizontal="left"), font=HEADER_FONT)

    current_row = 3
    if context["commissions"]:
        for item in context["commissions"]:
            current_row += 1
            row_date = _parse_row_date(item.get("DateCommission"))
            values = [
                _format_date(row_date) if row_date is not None else _safe_text(item.get("DateCommission")),
                _safe_text(item.get("Nom")),
                normalize_status_form_label(item.get("Statut")),
                int(item.get("NombreBacs", 0) or 0),
                float(item.get("MontantPaye", 0) or 0),
                float(item.get("Commissions", 0) or 0),
                float(item.get("Dettes", 0) or 0),
                float(item.get("NetAPayer", 0) or 0),
            ]
            for col_index, value in enumerate(values, start=1):
                _apply_cell_style(sheet.cell(current_row, col_index, value), alignment=Alignment(horizontal="left"))
            for col in range(5, 9):
                sheet.cell(current_row, col).number_format = MONEY_FORMAT

        totals_row = current_row + 1
        sheet.cell(totals_row, 1, "Totaux")
        sheet.cell(totals_row, 4, f"=SUM(D4:D{current_row})")
        sheet.cell(totals_row, 5, f"=SUM(E4:E{current_row})")
        sheet.cell(totals_row, 6, f"=SUM(F4:F{current_row})")
        sheet.cell(totals_row, 7, f"=SUM(G4:G{current_row})")
        sheet.cell(totals_row, 8, f"=SUM(H4:H{current_row})")
        for col_index in range(1, 9):
            _apply_cell_style(sheet.cell(totals_row, col_index), fill=SECTION_FILL, alignment=Alignment(horizontal="left"), font=SECTION_FONT if col_index == 1 else BODY_FONT)
        for col in range(5, 9):
            sheet.cell(totals_row, col).number_format = MONEY_FORMAT
        _add_table(sheet, 3, current_row, 8, "CommissionsPeriode")
    else:
        _apply_cell_style(sheet["A4"], alignment=Alignment(horizontal="left"), font=NOTE_FONT)
        sheet["A4"] = "Aucune commission enregistrée pour cette période."

    _autofit_columns(sheet)


def create_prevision_excel_workbook(
    target_date: date,
    destination: str | Path | None = None,
    generated_by: str = "",
    generated_role: str = "",
) -> Path:
    report_path = Path(destination) if destination is not None else DatabaseHelper.build_report_path(
        f"fiches-prevision-production-{target_date.strftime('%Y%m%d')}"
    )
    if report_path.suffix.lower() != ".xlsx":
        report_path = report_path.with_suffix(".xlsx")
    report_path.parent.mkdir(parents=True, exist_ok=True)

    rows = DatabaseHelper.list_previsions_by_date(target_date)
    summary = DatabaseHelper.get_prevision_summary_for_date(target_date)
    depositary_rows = [row for row in rows if str(row.get("Statut") or "").strip() == "Dépositaire"]
    mama_rows = [row for row in rows if str(row.get("Statut") or "").strip() == "Maman"]
    workbook = Workbook()

    try:
        _write_prevision_summary_sheet(
            workbook,
            target_date=target_date,
            summary=summary,
            generated_by=generated_by,
            generated_role=generated_role,
        )
        _write_prevision_order_sheet(
            workbook,
            title="Dépositaires",
            target_date=target_date,
            rows=depositary_rows,
            include_location=True,
            generated_by=generated_by,
            generated_role=generated_role,
        )
        _write_prevision_order_sheet(
            workbook,
            title="Mamans",
            target_date=target_date,
            rows=mama_rows,
            include_location=False,
            generated_by=generated_by,
            generated_role=generated_role,
        )
        workbook.save(report_path)
    except Exception as exc:
        raise ReportGenerationError("Impossible de générer les fiches de prévision Excel.") from exc
    finally:
        workbook.close()

    return report_path


def create_daily_excel_report(
    target_date: date,
    destination: str | Path | None = None,
    role: str = "Admin",
    generated_by: str = "",
    generated_role: str | None = None,
) -> Path:
    report_path = Path(destination) if destination is not None else DatabaseHelper.build_report_path(
        f"rapport-excel-journalier-{target_date.strftime('%Y%m%d')}"
    )
    if report_path.suffix.lower() != ".xlsx":
        report_path = report_path.with_suffix(".xlsx")
    report_path.parent.mkdir(parents=True, exist_ok=True)

    context = _build_report_context(target_date, role, generated_by=generated_by, generated_role=generated_role)
    workbook = Workbook()

    try:
        _build_summary_sheet(workbook, context)
        if "stock" in context["allowed_sections"]:
            _build_stock_sheet(workbook, context)
        if "production" in context["allowed_sections"]:
            _build_production_sheet(workbook, context)
        if "orders" in context["allowed_sections"]:
            _build_orders_sheet(workbook, context)
        if "cash" in context["allowed_sections"]:
            _build_cash_sheet(workbook, context)
        if "commissions" in context["allowed_sections"]:
            _build_commissions_sheet(workbook, context)
        if "workers" in context["allowed_sections"]:
            _build_payroll_sheet(workbook, context, "Travailleurs et paies du jour", "PaiesJour", include_date=False)
        workbook.save(report_path)
    except Exception as exc:
        raise ReportGenerationError("Impossible de générer le rapport Excel.") from exc
    finally:
        workbook.close()

    return report_path


def create_monthly_excel_report(
    target_date: date,
    destination: str | Path | None = None,
    role: str = "Admin",
    generated_by: str = "",
    generated_role: str | None = None,
) -> Path:
    report_path = Path(destination) if destination is not None else DatabaseHelper.build_report_path(
        f"rapport-excel-mensuel-{target_date.strftime('%Y%m')}"
    )
    if report_path.suffix.lower() != ".xlsx":
        report_path = report_path.with_suffix(".xlsx")
    report_path.parent.mkdir(parents=True, exist_ok=True)

    context = _build_monthly_report_context(
        target_date,
        role,
        generated_by=generated_by,
        generated_role=generated_role,
    )
    workbook = Workbook()

    try:
        _build_monthly_summary_sheet(workbook, context)
        if "stock" in context["allowed_sections"]:
            _build_monthly_stock_sheet(workbook, context)
        if "production" in context["allowed_sections"]:
            _build_production_period_sheet(workbook, context, "Production du mois", "ProductionMensuelle")
        if "orders" in context["allowed_sections"]:
            _build_monthly_orders_sheet(workbook, context)
        if "cash" in context["allowed_sections"]:
            _build_monthly_cash_sheet(workbook, context)
        if "commissions" in context["allowed_sections"]:
            _build_monthly_commissions_sheet(workbook, context)
        if "workers" in context["allowed_sections"]:
            _build_payroll_sheet(workbook, context, "Travailleurs et paies du mois", "PaiesMensuelles")
        workbook.save(report_path)
    except Exception as exc:
        raise ReportGenerationError("Impossible de générer le rapport Excel mensuel.") from exc
    finally:
        workbook.close()

    return report_path


def create_period_excel_report(
    start_date: date,
    end_date: date,
    destination: str | Path | None = None,
    role: str = "Admin",
    generated_by: str = "",
    generated_role: str | None = None,
) -> Path:
    start_date, end_date = _normalize_period_bounds(start_date, end_date)
    report_path = Path(destination) if destination is not None else DatabaseHelper.build_report_path(
        f"rapport-excel-periode-{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}"
    )
    if report_path.suffix.lower() != ".xlsx":
        report_path = report_path.with_suffix(".xlsx")
    report_path.parent.mkdir(parents=True, exist_ok=True)

    context = _build_period_report_context(
        start_date,
        end_date,
        role,
        generated_by=generated_by,
        generated_role=generated_role,
    )
    workbook = Workbook()

    try:
        _build_period_summary_sheet(workbook, context)
        if "stock" in context["allowed_sections"]:
            _build_period_stock_sheet(workbook, context)
        if "production" in context["allowed_sections"]:
            _build_production_period_sheet(workbook, context, "Production sur la période", "ProductionPeriode")
        if "orders" in context["allowed_sections"]:
            _build_period_orders_sheet(workbook, context)
        if "cash" in context["allowed_sections"]:
            _build_period_cash_sheet(workbook, context)
        if "commissions" in context["allowed_sections"]:
            _build_period_commissions_sheet(workbook, context)
        if "workers" in context["allowed_sections"]:
            _build_payroll_sheet(workbook, context, "Travailleurs et paies sur la période", "PaiesPeriode")
        workbook.save(report_path)
    except Exception as exc:
        raise ReportGenerationError("Impossible de générer le rapport Excel sur la période demandée.") from exc
    finally:
        workbook.close()

    return report_path
