from __future__ import annotations

from calendar import monthrange
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.shapes import Drawing
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .database import DatabaseHelper
from .report_branding import PDF_FONT_BOLD, PDF_FONT_REGULAR, REPORT_BLUE, REPORT_RED, register_pdf_fonts
from .reports import ReportGenerationError
from .version import APP_NAME


CASH_REPORT_ROLES = {"Admin", "Caissier"}
MONEY_FORMAT = '#,##0 "FC"'
THIN_BORDER = Border(
    left=Side(style="thin", color="AEBFD0"),
    right=Side(style="thin", color="AEBFD0"),
    top=Side(style="thin", color="AEBFD0"),
    bottom=Side(style="thin", color="AEBFD0"),
)
HEADER_FILL = PatternFill("solid", fgColor="DCE8F4")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")


def week_bounds(reference_date: date) -> tuple[date, date]:
    start = reference_date - timedelta(days=reference_date.weekday())
    return start, start + timedelta(days=6)


def month_bounds(reference_date: date) -> tuple[date, date]:
    return reference_date.replace(day=1), reference_date.replace(day=monthrange(reference_date.year, reference_date.month)[1])


def _format_fc(value: float) -> str:
    return f"{value:,.0f}".replace(",", " ") + " FC"


def _format_date(value: date) -> str:
    return value.strftime("%d/%m/%Y")


def _parse_row_date(value: Any) -> date | None:
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(str(value), "%Y-%m-%d").date()
    except ValueError:
        return None


def _ensure_cash_report_role(role: str) -> None:
    if role not in CASH_REPORT_ROLES:
        raise ReportGenerationError("Seuls l'administrateur et le caissier peuvent générer le bilan de caisse.")


def _cash_rows(start_date: date, end_date: date) -> list[dict[str, Any]]:
    return DatabaseHelper.list_cash_balance_by_period(start_date, end_date)


def _cash_totals(rows: list[dict[str, Any]]) -> dict[str, float]:
    balances = [float(row.get("Solde", 0) or 0) for row in rows]
    return {
        "received": sum(float(row.get("MontantRecu", 0) or 0) for row in rows),
        "paid_debts": sum(float(row.get("DettesPayeesAujourdHui", 0) or 0) for row in rows),
        "entries": sum(float(row.get("TotalEntrees", 0) or 0) for row in rows),
        "expenses": sum(float(row.get("MontantTotalDepenses", 0) or 0) for row in rows),
        "balance": sum(balances),
        "average_balance": sum(balances) / len(balances) if balances else 0.0,
        "positive_days": float(sum(1 for value in balances if value >= 0)),
        "negative_days": float(sum(1 for value in balances if value < 0)),
        "remaining_debts": float(rows[-1].get("DettesAccumuleesRestantes", 0) or 0) if rows else 0.0,
    }


def _payroll_rows(start_date: date, end_date: date, role: str) -> list[dict[str, Any]]:
    if role != "Admin":
        return []
    return DatabaseHelper.list_payrolls(start_date=start_date, end_date=end_date)


def _payroll_total(rows: list[dict[str, Any]]) -> float:
    return sum(float(row.get("MontantNet", 0) or 0) for row in rows)


def _payroll_totals_by_date(rows: list[dict[str, Any]]) -> dict[str, float]:
    totals: dict[str, float] = {}
    for row in rows:
        date_key = str(row.get("DatePaie", "") or "").strip()
        if not date_key:
            continue
        totals[date_key] = totals.get(date_key, 0.0) + float(row.get("MontantNet", 0) or 0)
    return totals


def _trend_text(rows: list[dict[str, Any]]) -> str:
    if len(rows) < 2:
        return "Pas assez de journées pour mesurer une tendance fiable."
    first = float(rows[0].get("Solde", 0) or 0)
    last = float(rows[-1].get("Solde", 0) or 0)
    if last > first:
        return "La courbe est montante : le solde final est supérieur au solde du début de période."
    if last < first:
        return "La courbe est descendante : le solde final est inférieur au solde du début de période."
    return "La courbe est stable sur la période observée."


def _pdf_styles() -> dict[str, ParagraphStyle]:
    register_pdf_fonts()
    sample = getSampleStyleSheet()
    return {
        "title": ParagraphStyle(
            "CashTitle",
            parent=sample["Heading1"],
            fontName=PDF_FONT_BOLD,
            fontSize=20,
            leading=24,
            alignment=TA_CENTER,
            textColor=colors.HexColor(REPORT_RED),
        ),
        "subtitle": ParagraphStyle(
            "CashSubtitle",
            parent=sample["Heading2"],
            fontName=PDF_FONT_BOLD,
            fontSize=13,
            leading=16,
            alignment=TA_CENTER,
            textColor=colors.HexColor(REPORT_BLUE),
        ),
        "body": ParagraphStyle(
            "CashBody",
            parent=sample["BodyText"],
            fontName=PDF_FONT_REGULAR,
            fontSize=10,
            leading=13,
        ),
        "section": ParagraphStyle(
            "CashSection",
            parent=sample["Heading2"],
            fontName=PDF_FONT_BOLD,
            fontSize=13,
            leading=16,
            textColor=colors.HexColor(REPORT_BLUE),
            spaceBefore=8,
            spaceAfter=5,
        ),
    }


def _pdf_table(rows: list[list[Any]], widths: list[float]) -> Table:
    table = Table(rows, colWidths=widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, 0), PDF_FONT_BOLD),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#DCE8F4")),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#AEBFD0")),
                ("FONTNAME", (0, 1), (-1, -1), PDF_FONT_REGULAR),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    return table


def _cash_chart(rows: list[dict[str, Any]], payrolls_by_date: dict[str, float] | None = None) -> Drawing:
    payrolls_by_date = payrolls_by_date or {}
    drawing = Drawing(170 * mm, 68 * mm)
    chart = VerticalBarChart()
    chart.x = 12
    chart.y = 24
    chart.height = 142
    chart.width = 430
    payroll_series = [
        payrolls_by_date.get(str(row.get("DateCaisse", "") or "").strip(), 0.0)
        for row in rows
    ]
    balance_after_payrolls = [
        float(row.get("Solde", 0) or 0) - payroll_series[index]
        for index, row in enumerate(rows)
    ]
    chart.data = [
        [float(row.get("TotalEntrees", 0) or 0) for row in rows],
        [float(row.get("MontantTotalDepenses", 0) or 0) for row in rows],
        payroll_series if payrolls_by_date else [float(row.get("Solde", 0) or 0) for row in rows],
        balance_after_payrolls if payrolls_by_date else [],
    ]
    chart.data = [series for series in chart.data if series]
    chart.categoryAxis.categoryNames = [
        (_format_date(row_date)[0:5] if (row_date := _parse_row_date(row.get("DateCaisse"))) else str(index + 1))
        for index, row in enumerate(rows)
    ]
    chart.bars[0].fillColor = colors.HexColor("#1E7D32")
    chart.bars[1].fillColor = colors.HexColor("#C27A00")
    chart.bars[2].fillColor = colors.HexColor("#6A3D9A" if payrolls_by_date else "#B30000")
    if payrolls_by_date and len(chart.bars) > 3:
        chart.bars[3].fillColor = colors.HexColor("#B30000")
    chart.valueAxis.valueMin = min(0, *(value for series in chart.data for value in series))
    chart.valueAxis.valueMax = max(1, *(value for series in chart.data for value in series))
    chart.valueAxis.labels.fontName = PDF_FONT_REGULAR
    chart.categoryAxis.labels.fontName = PDF_FONT_REGULAR
    drawing.add(chart)
    return drawing


def create_cash_balance_pdf_report(
    start_date: date,
    end_date: date,
    destination: str | Path,
    *,
    role: str,
    generated_by: str,
    generated_role: str,
    title: str,
) -> Path:
    _ensure_cash_report_role(role)
    rows = _cash_rows(start_date, end_date)
    payrolls = _payroll_rows(start_date, end_date, role)
    show_payrolls = role == "Admin"
    payroll_total = _payroll_total(payrolls)
    payrolls_by_date = _payroll_totals_by_date(payrolls)
    totals = _cash_totals(rows)
    balance_after_payrolls = totals["balance"] - payroll_total
    report_path = Path(destination)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    styles = _pdf_styles()

    doc = SimpleDocTemplate(
        str(report_path),
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=14 * mm,
        bottomMargin=16 * mm,
        title=f"{APP_NAME} - {title}",
        author="Kay Box Store",
    )

    elements: list[Any] = [
        Paragraph("BOULANGERIE LOMOTO", styles["title"]),
        Paragraph(title, styles["subtitle"]),
        Paragraph(f"Période : {_format_date(start_date)} au {_format_date(end_date)}", styles["subtitle"]),
        Paragraph(f"Généré par : {generated_by} ({generated_role})", styles["body"]),
        Spacer(1, 5 * mm),
        Paragraph("Résumé comptable", styles["section"]),
        _pdf_table(
            [
                ["Indicateur", "Valeur"],
                ["Montant reçu", _format_fc(totals["received"])],
                ["Dettes payées", _format_fc(totals["paid_debts"])],
                ["Total des entrées", _format_fc(totals["entries"])],
                ["Dépenses", _format_fc(totals["expenses"])],
                *( [["Paies travailleurs", _format_fc(payroll_total)], ["Balance après paies", _format_fc(balance_after_payrolls)]] if show_payrolls else [] ),
                ["Balance de la période", _format_fc(totals["balance"])],
                ["Solde moyen par jour", _format_fc(totals["average_balance"])],
                ["Dettes accumulées restantes", _format_fc(totals["remaining_debts"])],
            ],
            [72 * mm, 88 * mm],
        ),
        Spacer(1, 5 * mm),
        Paragraph("Évolution graphique", styles["section"]),
        _cash_chart(rows, payrolls_by_date if show_payrolls else None),
        Spacer(1, 5 * mm),
        Paragraph("Analyse", styles["section"]),
        Paragraph(
            f"{_trend_text(rows)} Jours positifs : {int(totals['positive_days'])}. "
            f"Jours négatifs : {int(totals['negative_days'])}. "
            "Une bonne direction se confirme lorsque les entrées progressent, que les dépenses restent maîtrisées "
            "et que les dettes accumulées diminuent.",
            styles["body"],
        ),
        Spacer(1, 5 * mm),
        Paragraph("Détail journalier", styles["section"]),
    ]
    detail_rows = (
        [["Date", "Reçu", "Dettes payées", "Entrées", "Dépenses", "Paies", "Solde réel", "Dettes restantes"]]
        if show_payrolls
        else [["Date", "Reçu", "Dettes payées", "Entrées", "Dépenses", "Solde", "Dettes restantes"]]
    )
    for row in rows:
        row_date = _parse_row_date(row.get("DateCaisse"))
        date_key = str(row.get("DateCaisse", "") or "").strip()
        base_values = [
            _format_date(row_date) if row_date else date_key,
            _format_fc(float(row.get("MontantRecu", 0) or 0)),
            _format_fc(float(row.get("DettesPayeesAujourdHui", 0) or 0)),
            _format_fc(float(row.get("TotalEntrees", 0) or 0)),
            _format_fc(float(row.get("MontantTotalDepenses", 0) or 0)),
        ]
        if show_payrolls:
            day_payroll = payrolls_by_date.get(date_key, 0.0)
            base_values.extend(
                [
                    _format_fc(day_payroll),
                    _format_fc(float(row.get("Solde", 0) or 0) - day_payroll),
                    _format_fc(float(row.get("DettesAccumuleesRestantes", 0) or 0)),
                ]
            )
        else:
            base_values.extend(
                [
                    _format_fc(float(row.get("Solde", 0) or 0)),
                    _format_fc(float(row.get("DettesAccumuleesRestantes", 0) or 0)),
                ]
            )
        detail_rows.append(base_values)
    detail_widths = [20 * mm, 23 * mm, 25 * mm, 23 * mm, 23 * mm, 22 * mm, 22 * mm, 22 * mm] if show_payrolls else [21 * mm, 25 * mm, 27 * mm, 25 * mm, 25 * mm, 25 * mm, 28 * mm]
    elements.append(_pdf_table(detail_rows, detail_widths))
    if show_payrolls:
        elements.extend(
            [
                Spacer(1, 5 * mm),
                Paragraph("Détail des paies des travailleurs", styles["section"]),
            ]
        )
        payroll_rows = [["Date", "Travailleur", "Période", "Brut", "Prime", "Avance", "Retenue", "Net"]]
        if payrolls:
            for row in payrolls:
                row_date = _parse_row_date(row.get("DatePaie"))
                payroll_rows.append(
                    [
                        _format_date(row_date) if row_date else str(row.get("DatePaie", "")),
                        str(row.get("NomComplet", "") or ""),
                        str(row.get("Periode", "") or ""),
                        _format_fc(float(row.get("MontantBrut", 0) or 0)),
                        _format_fc(float(row.get("Prime", 0) or 0)),
                        _format_fc(float(row.get("Avance", 0) or 0)),
                        _format_fc(float(row.get("Retenue", 0) or 0)),
                        _format_fc(float(row.get("MontantNet", 0) or 0)),
                    ]
                )
        else:
            payroll_rows.append(
                ["-", "Aucune paie enregistrée", "-", _format_fc(0), _format_fc(0), _format_fc(0), _format_fc(0), _format_fc(0)]
            )
        elements.append(_pdf_table(payroll_rows, [18 * mm, 38 * mm, 18 * mm, 21 * mm, 17 * mm, 17 * mm, 17 * mm, 21 * mm]))
    doc.build(elements)
    return report_path


def create_cash_balance_excel_report(
    start_date: date,
    end_date: date,
    destination: str | Path,
    *,
    role: str,
    generated_by: str,
    generated_role: str,
    title: str,
) -> Path:
    _ensure_cash_report_role(role)
    rows = _cash_rows(start_date, end_date)
    payrolls = _payroll_rows(start_date, end_date, role)
    show_payrolls = role == "Admin"
    payroll_total = _payroll_total(payrolls)
    payrolls_by_date = _payroll_totals_by_date(payrolls)
    totals = _cash_totals(rows)
    balance_after_payrolls = totals["balance"] - payroll_total
    report_path = Path(destination)
    report_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Bilan caisse"
    sheet["A1"] = "BOULANGERIE LOMOTO"
    sheet["A2"] = title
    sheet["A3"] = f"Période : {_format_date(start_date)} au {_format_date(end_date)}"
    sheet["A4"] = f"Généré par : {generated_by} ({generated_role})"
    for row_index in range(1, 5):
        sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=7)
        cell = sheet.cell(row_index, 1)
        cell.font = Font(name="Poppins", size=16 if row_index == 1 else 12, bold=True, color="B30000" if row_index == 1 else "1F4E78")
        cell.alignment = Alignment(horizontal="center")

    summary_rows = [
        ("Montant reçu", totals["received"]),
        ("Dettes payées", totals["paid_debts"]),
        ("Total des entrées", totals["entries"]),
        ("Dépenses", totals["expenses"]),
        *( [("Paies travailleurs", payroll_total), ("Balance après paies", balance_after_payrolls)] if show_payrolls else [] ),
        ("Balance de la période", totals["balance"]),
        ("Solde moyen par jour", totals["average_balance"]),
        ("Dettes accumulées restantes", totals["remaining_debts"]),
    ]
    sheet.cell(6, 1, "Indicateur")
    sheet.cell(6, 2, "Valeur")
    for col in range(1, 3):
        cell = sheet.cell(6, col)
        cell.fill = HEADER_FILL
        cell.font = Font(name="Poppins", bold=True)
        cell.border = THIN_BORDER
    for index, (label, value) in enumerate(summary_rows, start=7):
        sheet.cell(index, 1, label)
        sheet.cell(index, 2, value)
        sheet.cell(index, 2).number_format = MONEY_FORMAT
        for col in range(1, 3):
            sheet.cell(index, col).border = THIN_BORDER
            sheet.cell(index, col).font = Font(name="Poppins")

    data_start = 17
    headers = ["Date", "Bacs", "Attendu", "Reçu", "Dettes payées", "Entrées", "Dépenses"]
    if show_payrolls:
        headers.extend(["Paies travailleurs", "Solde après paies"])
    headers.extend(["Solde", "Solde cumulé", "Dettes restantes"])
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(data_start, col, header)
        cell.fill = HEADER_FILL
        cell.font = Font(name="Poppins", bold=True)
        cell.border = THIN_BORDER
    for row_index, row in enumerate(rows, start=data_start + 1):
        row_date = _parse_row_date(row.get("DateCaisse"))
        date_key = str(row.get("DateCaisse", "") or "").strip()
        values = [
            _format_date(row_date) if row_date else str(row.get("DateCaisse", "")),
            int(row.get("NombreTotalBacs", 0) or 0),
            float(row.get("MontantAttendu", 0) or 0),
            float(row.get("MontantRecu", 0) or 0),
            float(row.get("DettesPayeesAujourdHui", 0) or 0),
            float(row.get("TotalEntrees", 0) or 0),
            float(row.get("MontantTotalDepenses", 0) or 0),
        ]
        if show_payrolls:
            day_payroll = payrolls_by_date.get(date_key, 0.0)
            values.extend([day_payroll, float(row.get("Solde", 0) or 0) - day_payroll])
        values.extend(
            [
                float(row.get("Solde", 0) or 0),
                float(row.get("SoldeCumule", 0) or 0),
                float(row.get("DettesAccumuleesRestantes", 0) or 0),
            ]
        )
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, col, value)
            cell.border = THIN_BORDER
            cell.font = Font(name="Poppins")
            if col >= 3:
                cell.number_format = MONEY_FORMAT

    last_row = data_start + len(rows)
    if rows:
        bar = BarChart()
        bar.title = "Entrées, dépenses, paies et solde" if show_payrolls else "Entrées, dépenses et solde"
        bar.y_axis.title = "Montant"
        bar.x_axis.title = "Date"
        bar_last_col = 9 if show_payrolls else 8
        bar.add_data(Reference(sheet, min_col=6, max_col=bar_last_col, min_row=data_start, max_row=last_row), titles_from_data=True)
        bar.set_categories(Reference(sheet, min_col=1, min_row=data_start + 1, max_row=last_row))
        bar.height = 8
        bar.width = 18
        sheet.add_chart(bar, "L6")

        line = LineChart()
        line.title = "Solde cumulé"
        line.y_axis.title = "Montant"
        line.x_axis.title = "Date"
        cumulative_col = 11 if show_payrolls else 9
        line.add_data(Reference(sheet, min_col=cumulative_col, min_row=data_start, max_row=last_row), titles_from_data=True)
        line.set_categories(Reference(sheet, min_col=1, min_row=data_start + 1, max_row=last_row))
        line.height = 8
        line.width = 18
        sheet.add_chart(line, "L22")

    if show_payrolls:
        payroll_sheet = workbook.create_sheet("Paies travailleurs")
        payroll_sheet["A1"] = "Paies des travailleurs"
        payroll_sheet["A2"] = f"Période : {_format_date(start_date)} au {_format_date(end_date)}"
        for row_index in range(1, 3):
            payroll_sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=8)
            cell = payroll_sheet.cell(row_index, 1)
            cell.font = Font(name="Poppins", size=14 if row_index == 1 else 11, bold=True, color="B30000" if row_index == 1 else "1F4E78")
            cell.alignment = Alignment(horizontal="center")
        payroll_headers = ["Date", "Travailleur", "Période", "Brut", "Prime", "Avance", "Retenue", "Net"]
        for col, header in enumerate(payroll_headers, start=1):
            cell = payroll_sheet.cell(4, col, header)
            cell.fill = HEADER_FILL
            cell.font = Font(name="Poppins", bold=True)
            cell.border = THIN_BORDER
        current_row = 4
        if payrolls:
            for item in payrolls:
                current_row += 1
                row_date = _parse_row_date(item.get("DatePaie"))
                values = [
                    _format_date(row_date) if row_date else str(item.get("DatePaie", "")),
                    str(item.get("NomComplet", "") or ""),
                    str(item.get("Periode", "") or ""),
                    float(item.get("MontantBrut", 0) or 0),
                    float(item.get("Prime", 0) or 0),
                    float(item.get("Avance", 0) or 0),
                    float(item.get("Retenue", 0) or 0),
                    float(item.get("MontantNet", 0) or 0),
                ]
                for col, value in enumerate(values, start=1):
                    cell = payroll_sheet.cell(current_row, col, value)
                    cell.border = THIN_BORDER
                    cell.font = Font(name="Poppins")
                    if col >= 4:
                        cell.number_format = MONEY_FORMAT
            total_row = current_row + 1
            payroll_sheet.cell(total_row, 1, "Totaux")
            for col in range(4, 9):
                col_letter = get_column_letter(col)
                payroll_sheet.cell(total_row, col, f"=SUM({col_letter}5:{col_letter}{current_row})")
                payroll_sheet.cell(total_row, col).number_format = MONEY_FORMAT
            for col in range(1, 9):
                cell = payroll_sheet.cell(total_row, col)
                cell.fill = SECTION_FILL
                cell.border = THIN_BORDER
                cell.font = Font(name="Poppins", bold=(col == 1))
        else:
            payroll_sheet.cell(5, 1, "Aucune paie enregistrée pour cette période.")
            payroll_sheet.cell(5, 1).font = Font(name="Poppins", italic=True)
        for column_cells in payroll_sheet.columns:
            letter = get_column_letter(column_cells[0].column)
            payroll_sheet.column_dimensions[letter].width = min(max(max(len(str(cell.value or "")) for cell in column_cells) + 2, 12), 28)

    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        sheet.column_dimensions[letter].width = min(max(max(len(str(cell.value or "")) for cell in column_cells) + 2, 12), 28)

    try:
        workbook.save(report_path)
    finally:
        workbook.close()
    return report_path
