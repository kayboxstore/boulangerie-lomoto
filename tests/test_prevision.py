"""Tests de non-régression du module Prévisions.

Régression : ``PREVISION_STATUSES`` et ``PREVISION_LOCATIONS`` étaient utilisées
dans ``PrevisionWindow`` sans jamais être définies -> le module « Prévisions »
plantait (NameError) dès son ouverture.
"""

from __future__ import annotations

from datetime import date, timedelta
import json
from pathlib import Path

import pytest

from openpyxl import load_workbook

from boulangerie_app.app import ROLE_MODULE_ACCESS, ROLE_READ_ONLY_MODULES
from boulangerie_app.connected_server import _is_method_allowed_for_session
from boulangerie_app.database import DatabaseHelper
from boulangerie_app.excel_reports import create_prevision_excel_workbook
from boulangerie_web_pro.server import (
    ROLE_MODULES,
    WebProHandler,
    _can_write,
    _is_read_only,
)
from boulangerie_app.status_labels import DEPOSITARY_STATUS


def test_prevision_constants_are_defined():
    import boulangerie_app.app as app

    assert isinstance(app.PREVISION_STATUSES, (list, tuple)) and app.PREVISION_STATUSES
    assert isinstance(app.PREVISION_LOCATIONS, (list, tuple)) and app.PREVISION_LOCATIONS
    # Seuls Dépositaire et Maman sont des statuts de prévision valides.
    assert set(app.PREVISION_STATUSES) == {DEPOSITARY_STATUS, "Maman"}


def test_prevision_order_roundtrip(db):
    today = date.today()
    db.add_prevision_order(today, "Dépôt 1", "Client Test", DEPOSITARY_STATUS, 5, 0, 3, 0)
    rows = db.list_previsions_by_date(today)
    assert len(rows) == 1
    assert rows[0]["Client"] == "Client Test"
    assert rows[0]["Localisation"] == "Dépôt 1"


def test_future_prevision_is_allowed(db):
    tomorrow = date.today() + timedelta(days=1)
    db.add_prevision_order(tomorrow, "Dépôt 2", "Client demain", DEPOSITARY_STATUS, 2, 1, 0, 0)

    rows = db.list_previsions_by_date(tomorrow)

    assert len(rows) == 1
    assert rows[0]["DatePrevision"] == tomorrow.isoformat()
    assert rows[0]["TotalArticles"] == 3


def test_prevision_module_is_visible_for_operational_roles():
    assert "Prévisions" in ROLE_MODULE_ACCESS["Admin"]
    assert "Prévisions" in ROLE_MODULE_ACCESS["Directeur Général"]
    assert "Prévisions" in ROLE_MODULE_ACCESS["Chargé de la production"]
    assert "Prévisions" in ROLE_MODULE_ACCESS["Gestionnaire des commandes"]
    assert "Prévisions" in ROLE_READ_ONLY_MODULES["Directeur Général"]
    assert "Prévisions" not in ROLE_MODULE_ACCESS["Caissier"]


def test_connected_prevision_permissions_match_desktop_roles():
    production = {"role": "Chargé de la production", "identifiant": "prod"}
    orders = {"role": "Gestionnaire des commandes", "identifiant": "cmd"}
    director = {"role": "Directeur Général", "identifiant": "dg"}
    cashier = {"role": "Caissier", "identifiant": "caisse"}

    assert _is_method_allowed_for_session("add_prevision_order", [], production)
    assert _is_method_allowed_for_session("add_prevision_order", [], orders)
    assert _is_method_allowed_for_session("list_previsions", [], director)
    assert not _is_method_allowed_for_session("add_prevision_order", [], director)
    assert not _is_method_allowed_for_session("list_previsions", [], cashier)


def test_web_prevision_permissions_match_desktop_roles():
    assert "previsions" in ROLE_MODULES["Admin"]
    assert "previsions" in ROLE_MODULES["Directeur Général"]
    assert "previsions" in ROLE_MODULES["Chargé de la production"]
    assert "previsions" in ROLE_MODULES["Gestionnaire des commandes"]
    assert "previsions" not in ROLE_MODULES["Caissier"]
    assert _can_write("Admin", "previsions")
    assert _can_write("Chargé de la production", "previsions")
    assert _can_write("Gestionnaire des commandes", "previsions")
    assert not _can_write("Directeur Général", "previsions")
    assert _is_read_only("Directeur Général", "previsions")


def _web_handler_for(session):
    handler = object.__new__(WebProHandler)
    responses = []
    handler._require_license_active = lambda: None
    handler._require_session = lambda: session
    handler._send_json = lambda payload, *args, **kwargs: responses.append(payload)
    return handler, responses


def test_web_prevision_api_roundtrip_and_excel_export(db):
    tomorrow = date.today() + timedelta(days=1)
    session = {
        "identifiant": "admin.web",
        "fullName": "Admin Web",
        "role": "Admin",
    }
    handler, responses = _web_handler_for(session)

    handler._handle_api_post(
        "/api/previsions",
        {
            "date": tomorrow.isoformat(),
            "location": "Dépôt 1",
            "client": "Client Web",
            "status": DEPOSITARY_STATUS,
            "square1500": 2,
            "square1000": 1,
            "baguette500": 3,
            "baguette1000": 0,
        },
    )
    assert responses[-1]["ok"] is True

    handler._handle_api_get(
        "/api/previsions",
        {"date": [tomorrow.isoformat()], "all": ["0"]},
    )
    result = responses[-1]
    assert result["ok"] is True
    assert result["summary"]["TotalArticlesPrevus"] == 6
    assert result["rows"][0]["Client"] == "Client Web"
    record_id = result["rows"][0]["Id"]

    handler._handle_api_post(
        "/api/previsions",
        {
            "id": record_id,
            "date": tomorrow.isoformat(),
            "location": "Dépôt 2",
            "client": "Client Web modifié",
            "status": DEPOSITARY_STATUS,
            "square1500": 4,
            "square1000": 0,
            "baguette500": 0,
            "baguette1000": 0,
        },
    )
    assert db.list_previsions_by_date(tomorrow)[0]["Client"] == "Client Web modifié"

    handler._handle_api_post("/api/previsions/export", {"date": tomorrow.isoformat()})
    export = responses[-1]
    assert export["ok"] is True
    assert Path(export["path"]).is_file()
    assert export["url"].startswith("/api/reports/file?token=")

    handler._handle_api_delete("/api/previsions", {"id": [str(record_id)]})
    assert responses[-1]["ok"] is True
    assert db.list_previsions_by_date(tomorrow) == []


def test_web_prevision_api_rejects_unauthorized_roles(db):
    tomorrow = date.today() + timedelta(days=1)
    cashier, _ = _web_handler_for({"identifiant": "cash", "role": "Caissier"})
    with pytest.raises(PermissionError, match="non autorisé"):
        cashier._handle_api_get(
            "/api/previsions",
            {"date": [tomorrow.isoformat()], "all": ["0"]},
        )

    director, _ = _web_handler_for({"identifiant": "dg", "role": "Directeur Général"})
    with pytest.raises(PermissionError, match="ne peut pas le modifier"):
        director._handle_api_post(
            "/api/previsions",
            {
                "date": tomorrow.isoformat(),
                "client": "Lecture seule",
                "status": "Maman",
                "square1500": 1,
            },
        )


def test_prevision_screen_is_shared_with_android_wrapper():
    root = Path(__file__).resolve().parents[1]
    web_source = (root / "boulangerie_web_pro" / "static" / "app.js").read_text(encoding="utf-8")
    android_config = json.loads((root / "android-apk" / "capacitor.config.json").read_text(encoding="utf-8"))

    assert "previsions: previsionsView" in web_source
    assert 'data-allow-future="true"' in web_source
    assert "/api/previsions/export" in web_source
    assert android_config["server"]["url"] == "https://app.boulangerie-lomoto.com"


def test_prevision_excel_workbook_contains_three_printable_sheets_and_sanitizes_text(db, tmp_path):
    tomorrow = date.today() + timedelta(days=1)
    db.add_prevision_order(tomorrow, "Dépôt 1", "=2+3", DEPOSITARY_STATUS, 5, 0, 3, 0)
    db.add_prevision_order(tomorrow, "", "Maman Test", "Maman", 0, 2, 0, 1)
    destination = tmp_path / "previsions.xlsx"

    result = create_prevision_excel_workbook(
        tomorrow,
        destination,
        generated_by="Admin Test",
        generated_role="Admin",
    )

    assert result == destination
    workbook = load_workbook(destination, data_only=False)
    try:
        assert workbook.sheetnames == ["Résumé", "Dépositaires", "Mamans"]
        depositary_values = [cell.value for row in workbook["Dépositaires"].iter_rows() for cell in row]
        assert "'=2+3" in depositary_values
        assert "=2+3" not in depositary_values
        assert workbook["Dépositaires"].page_setup.orientation == "landscape"
        summary_values = [cell.value for row in workbook["Résumé"].iter_rows() for cell in row]
        assert 11 in summary_values
        assert 12000 in summary_values
    finally:
        workbook.close()
